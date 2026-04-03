"""
Invoice Extractor - PDF to Excel + Postgres JSON
==================================================
Extracts all data from multi-page repair invoices into:
  - Excel file (.xlsx)
  - JSON file for Postgres import via PipelineX API

Usage:
  python extract_invoice.py "invoice.pdf"                    # Excel only
  python extract_invoice.py "invoice.pdf" --json             # Excel + JSON
  python extract_invoice.py "invoice.pdf" --import           # Excel + push to PipelineX API
  python extract_invoice.py "invoice.pdf" --import --url http://localhost:3000  # Custom API URL
"""

import sys
import os
import re
import json
import pdfplumber
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill


def extract_invoice(pdf_path, output_path=None):
    from vendor_parsers import parse_invoice
    
    if not output_path:
        base = os.path.splitext(pdf_path)[0]
        output_path = f"{base}.xlsx"

    # Use universal vendor parser
    data = parse_invoice(pdf_path)
    
    # Extract fields from standardized dict
    date_open = data.get("invoice_date", "")
    written_by = data.get("written_by", "")
    terms = data.get("terms", "")
    po_num = data.get("po_number", "")
    unit_num = data.get("unit_number", "")
    plate_num = data.get("plate_number", "")
    year = data.get("year", "")
    make = data.get("make", "")
    model = data.get("model", "")
    mileage = data.get("mileage_hours", "")
    vin = data.get("vin", "")
    engine = data.get("engine", "")
    line_items = data.get("line_items", [])
    work_sections = data.get("work_descriptions", [])
    
    # Get totals from parsed data
    parts_match = data.get("parts_total")
    labor_match = data.get("labor_total")
    supplies_match = data.get("shop_supplies")
    subtotal_match = data.get("subtotal")
    tax_match = data.get("sales_tax")
    total_match = data.get("grand_total")
    
    # Extract customer info (not available in all vendors, use defaults)
    customer_block = ""
    ship_to_block = ""
    time_val = ""

    # ============================================
    # BUILD EXCEL FILE
    # ============================================
    wb = openpyxl.Workbook()

    # --- Styles ---
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="1A3A5C", end_color="1A3A5C", fill_type="solid")
    section_font = Font(bold=True, size=11, color="1A3A5C")
    money_format = '#,##0.00'
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # ============================================
    # SHEET 1: Invoice Summary
    # ============================================
    ws = wb.active
    ws.title = "Invoice Summary"
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 50

    ws.append(["INVOICE SUMMARY"])
    ws['A1'].font = Font(bold=True, size=14, color="1A3A5C")

    ws.append([])
    ws.append(["Date Opened", date_open])
    ws.append(["Written By", written_by])
    ws.append(["Terms", terms])
    ws.append(["PO #", po_num])
    ws.append([])
    ws.append(["CUSTOMER"])
    ws['A8'].font = section_font
    for line in customer_block.split("\n") if customer_block else []:
        ws.append(["", line.strip()])
    ws.append([])
    ws.append(["VEHICLE INFORMATION"])
    ws[f'A{ws.max_row}'].font = section_font
    ws.append(["Unit #", unit_num])
    ws.append(["Plate #", plate_num])
    ws.append(["Year", year])
    ws.append(["Make", make])
    ws.append(["Model", model])
    ws.append(["Mileage/Hours", mileage])
    ws.append(["VIN", vin])
    ws.append(["Engine", engine])

    ws.append([])
    ws.append(["TOTALS"])
    ws[f'A{ws.max_row}'].font = section_font
    if parts_match: ws.append(["Parts", parts_match])
    if labor_match: ws.append(["Labor", labor_match])
    if supplies_match: ws.append(["Shop Supplies", supplies_match])
    if subtotal_match: ws.append(["Sub Total", subtotal_match])
    if tax_match: ws.append(["Sales Tax", tax_match])
    if total_match: ws.append(["TOTAL", total_match])

    # Format money cells
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=2, max_col=2):
        for cell in row:
            if isinstance(cell.value, float):
                cell.number_format = money_format

    # ============================================
    # SHEET 2: Line Items (Parts)
    # ============================================
    ws2 = wb.create_sheet("Line Items")
    ws2.column_dimensions['A'].width = 8
    ws2.column_dimensions['B'].width = 10
    ws2.column_dimensions['C'].width = 50
    ws2.column_dimensions['D'].width = 12
    ws2.column_dimensions['E'].width = 12

    headers = ["Page", "Qty", "Description", "Unit Price", "Amount"]
    ws2.append(headers)
    for col_idx, header in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    for item in line_items:
        ws2.append([
            item.get("page", ""),
            item.get("qty", ""),
            item.get("description", ""),
            item.get("unit_price", ""),
            item.get("amount", ""),
        ])

    # Format money columns
    for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row, min_col=4, max_col=5):
        for cell in row:
            if isinstance(cell.value, float):
                cell.number_format = money_format
            cell.border = thin_border

    for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row, min_col=1, max_col=3):
        for cell in row:
            cell.border = thin_border

    # ============================================
    # SHEET 3: Work Descriptions
    # ============================================
    ws3 = wb.create_sheet("Work Descriptions")
    ws3.column_dimensions['A'].width = 8
    ws3.column_dimensions['B'].width = 100

    ws3.append(["Page", "Work Description"])
    ws3['A1'].font = header_font
    ws3['A1'].fill = header_fill
    ws3['B1'].font = header_font
    ws3['B1'].fill = header_fill

    for section in work_sections:
        ws3.append([section.get("page", ""), section.get("work_text", "")])
        ws3[f'B{ws3.max_row}'].alignment = Alignment(wrap_text=True)

    # ============================================
    # SHEET 4: Raw Data (summary from parsed data)
    # ============================================
    ws4 = wb.create_sheet("Raw Data")
    ws4['A1'] = "Parsed from vendor:"
    ws4['B1'] = data.get("vendor", "Unknown")
    ws4['A2'] = "Source file:"
    ws4['B2'] = data.get("source_file", "")

    # Save
    wb.save(output_path)
    print(f"Extracted to: {output_path}")
    print(f"  - Invoice Summary (header, vehicle, totals)")
    print(f"  - Line Items ({len(line_items)} parts/materials)")
    print(f"  - Work Descriptions ({len(work_sections)} sections)")
    print(f"  - Raw Data (all cells as-is)")


def extract_invoice_data(pdf_path):
    """Extract invoice data and return as a dict ready for Postgres import.
    Thin wrapper around parse_invoice() from vendor_parsers."""
    from vendor_parsers import parse_invoice
    
    # Use universal vendor parser
    data = parse_invoice(pdf_path)
    
    # The parser already returns the standardized dict we need
    return data


def export_json(pdf_path, json_path=None):
    """Extract invoice and save as JSON for Postgres import."""
    data = extract_invoice_data(pdf_path)
    if not json_path:
        json_path = os.path.splitext(pdf_path)[0] + ".json"
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2)
    print(f"JSON exported: {json_path}")
    print(f"  Vendor: {data['vendor']}")
    print(f"  Date: {data['invoice_date']}")
    print(f"  Unit: {data['unit_number']} ({data['year']} {data['make']} {data['model']})")
    print(f"  Line items: {len(data['line_items'])}")
    print(f"  Work sections: {len(data['work_descriptions'])}")
    print(f"  Total: ${data['grand_total']}")
    return data


def _generate_invoice_id(data):
    """Generate a short unique invoice ID from vendor + filename."""
    vendor_codes = {
        "Colorado Truck Repair": "CTR",
        "MHC Kenworth": "MHC",
        "Purcell Tire": "PUR",
        "Southern Tire Mart": "STM",
        "Rush Truck Centers": "RSH",
        "Randy's Towing": "RND",
        "NAPA": "NAP",
        "Bosselman": "BOS",
        "PSC Custom": "PSC",
        "Service Auto Glass": "SAG",
    }
    code = vendor_codes.get(data.get("vendor", ""), "INV")
    # Extract a number from the filename
    nums = re.findall(r'\d+', data.get("source_file", ""))
    num = nums[0] if nums else "000"
    return f"{code}-{num}"


def to_maintenance_tracker_rows(data):
    """Convert extracted invoice data to multiple maintenance_tracker rows —
    one row per line item. Each row carries the full invoice header so you
    can filter/pivot by vendor, truck, date, etc."""

    invoice_id = _generate_invoice_id(data)

    # Split work descriptions into requested vs completed
    work_requested = []
    work_completed = []
    for section in data.get("work_descriptions", []):
        text = section.get("work_text", "")
        if "Work Completed" in text:
            parts = text.split("Work Completed :")
            if len(parts) > 1:
                work_completed.append(parts[1].strip())
            req_parts = parts[0].split("Work Requested :")
            if len(req_parts) > 1:
                work_requested.append(req_parts[1].strip())
        elif "Work Requested" in text:
            parts = text.split("Work Requested :")
            if len(parts) > 1:
                work_requested.append(parts[1].strip())
        elif "COMPLAINT" in text.upper() or "CORRECTION" in text.upper():
            work_completed.append(text.strip())
        elif text.strip():
            work_completed.append(text.strip())

    work_req_str = "\n---\n".join(work_requested)
    work_comp_str = "\n---\n".join(work_completed)

    asset_name = f"{data.get('year', '')} {data.get('make', '')} {data.get('model', '')}".strip()

    rows = []
    line_items = data.get("line_items", [])

    if not line_items:
        # No line items — still create one row for the invoice
        line_items = [{"qty": 0, "description": "Service/Labor", "unit_price": 0,
                       "amount": data.get("grand_total", 0), "page": 1}]

    for i, item in enumerate(line_items):
        line_id = f"{invoice_id}-{str(i+1).zfill(3)}"
        rows.append({
            "Maintenance ID": invoice_id,
            "Line Item ID": line_id,
            "Asset ID": data.get("unit_number", ""),
            "Asset Name": asset_name,
            "Service Type": "Repair",
            "Completed Date": data.get("invoice_date", ""),
            "Status": "Completed",
            "Vendor": data.get("vendor", ""),
            "PO Number": data.get("po_number", ""),
            "Unit Number": data.get("unit_number", ""),
            "Plate Number": data.get("plate_number", ""),
            "Year": data.get("year", ""),
            "Make": data.get("make", ""),
            "Model": data.get("model", ""),
            "VIN": data.get("vin", ""),
            "Engine": data.get("engine", ""),
            "Mileage Hours": data.get("mileage_hours", ""),
            "Written By": data.get("written_by", ""),
            "Qty": item.get("qty", 0),
            "Item Description": item.get("description", ""),
            "Unit Price": item.get("unit_price", 0),
            "Line Amount": item.get("amount", 0),
            "Parts Cost": data.get("parts_total", 0),
            "Labor Cost": data.get("labor_total", 0),
            "Shop Supplies": data.get("shop_supplies", 0),
            "Sales Tax": data.get("sales_tax", 0),
            "Invoice Total": data.get("grand_total", 0),
            "Work Requested": work_req_str,
            "Work Completed": work_comp_str,
            "Source File": data.get("source_file", ""),
        })

    return rows


def export_maintenance_xlsx(pdf_paths, output_path="maintenance_import.xlsx"):
    """Convert invoice PDFs to a two-sheet XLSX:
      Sheet 1: Invoices — one row per invoice with totals, vehicle, work descriptions
      Sheet 2: Line Items — one row per part/service, linked by Invoice ID
    Uses the universal vendor parser for auto-detection."""
    from vendor_parsers import parse_invoice

    wb = openpyxl.Workbook()

    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="1A3A5C", end_color="1A3A5C", fill_type="solid")
    money_format = '#,##0.00'

    def style_headers(ws, headers):
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

    # ============================================
    # SHEET 1: INVOICES
    # ============================================
    ws_inv = wb.active
    ws_inv.title = "Invoices"

    inv_headers = [
        "Invoice ID", "Completed Date", "Vendor", "PO Number",
        "Unit Number", "Asset Name", "Year", "Make", "Model",
        "Plate Number", "VIN", "Engine", "Mileage Hours", "Written By",
        "Parts Cost", "Labor Cost", "Shop Supplies", "Sales Tax", "Invoice Total",
        "Line Item Count",
        "Work Requested", "Work Completed", "Source File",
    ]
    ws_inv.append(inv_headers)
    style_headers(ws_inv, inv_headers)

    # ============================================
    # SHEET 2: LINE ITEMS
    # ============================================
    ws_items = wb.create_sheet("Line Items")

    item_headers = [
        "Line Item ID", "Invoice ID", "Completed Date", "Vendor",
        "Unit Number", "Asset Name", "VIN",
        "Qty", "Item Description", "Unit Price", "Line Amount",
    ]
    ws_items.append(item_headers)
    style_headers(ws_items, item_headers)

    total_invoices = 0
    total_items = 0

    for pdf_path in pdf_paths:
        print(f"Processing: {os.path.basename(pdf_path)}")
        try:
            data = parse_invoice(pdf_path)
            invoice_id = _generate_invoice_id(data)

            # Work descriptions
            work_requested = []
            work_completed = []
            for section in data.get("work_descriptions", []):
                text = section.get("work_text", "")
                if "Work Completed" in text:
                    parts = text.split("Work Completed :")
                    if len(parts) > 1:
                        work_completed.append(parts[1].strip())
                    req_parts = parts[0].split("Work Requested :")
                    if len(req_parts) > 1:
                        work_requested.append(req_parts[1].strip())
                elif "Work Requested" in text:
                    parts = text.split("Work Requested :")
                    if len(parts) > 1:
                        work_requested.append(parts[1].strip())
                elif text.strip():
                    work_completed.append(text.strip())

            asset_name = f"{data.get('year', '')} {data.get('make', '')} {data.get('model', '')}".strip()
            line_items = data.get("line_items", [])

            # --- Invoice row ---
            ws_inv.append([
                invoice_id,
                data.get("invoice_date", ""),
                data.get("vendor", ""),
                data.get("po_number", ""),
                data.get("unit_number", ""),
                asset_name,
                data.get("year", ""),
                data.get("make", ""),
                data.get("model", ""),
                data.get("plate_number", ""),
                data.get("vin", ""),
                data.get("engine", ""),
                data.get("mileage_hours", ""),
                data.get("written_by", ""),
                data.get("parts_total", 0),
                data.get("labor_total", 0),
                data.get("shop_supplies", 0),
                data.get("sales_tax", 0),
                data.get("grand_total", 0),
                len(line_items),
                "\n---\n".join(work_requested),
                "\n---\n".join(work_completed),
                data.get("source_file", ""),
            ])
            total_invoices += 1

            # --- Line item rows ---
            for i, item in enumerate(line_items):
                line_id = f"{invoice_id}-{str(i+1).zfill(3)}"
                ws_items.append([
                    line_id,
                    invoice_id,
                    data.get("invoice_date", ""),
                    data.get("vendor", ""),
                    data.get("unit_number", ""),
                    asset_name,
                    data.get("vin", ""),
                    item.get("qty", 0),
                    item.get("description", ""),
                    item.get("unit_price", 0),
                    item.get("amount", 0),
                ])
                total_items += 1

            print(f"  -> {data['vendor']} | Unit: {data.get('unit_number', '?')} | ${data.get('grand_total', 0)} | {len(line_items)} items")
        except Exception as e:
            print(f"  ERROR: {e}")

    # ============================================
    # FORMAT COLUMNS
    # ============================================
    # Invoices sheet
    inv_money = ["Parts Cost", "Labor Cost", "Shop Supplies", "Sales Tax", "Invoice Total"]
    for col_idx, header in enumerate(inv_headers, 1):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        ws_inv.column_dimensions[col_letter].width = max(len(header) + 2, 14)
        if header in inv_money:
            for row_idx in range(2, total_invoices + 2):
                cell = ws_inv.cell(row=row_idx, column=col_idx)
                if isinstance(cell.value, (int, float)):
                    cell.number_format = money_format
    for header in ["Work Requested", "Work Completed", "Asset Name"]:
        if header in inv_headers:
            col_letter = openpyxl.utils.get_column_letter(inv_headers.index(header) + 1)
            ws_inv.column_dimensions[col_letter].width = 60

    # Line Items sheet
    item_money = ["Unit Price", "Line Amount"]
    for col_idx, header in enumerate(item_headers, 1):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        ws_items.column_dimensions[col_letter].width = max(len(header) + 2, 14)
        if header in item_money:
            for row_idx in range(2, total_items + 2):
                cell = ws_items.cell(row=row_idx, column=col_idx)
                if isinstance(cell.value, (int, float)):
                    cell.number_format = money_format
    if "Item Description" in item_headers:
        col_letter = openpyxl.utils.get_column_letter(item_headers.index("Item Description") + 1)
        ws_items.column_dimensions[col_letter].width = 55

    # Add borders to both sheets
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    for ws, max_row, max_col in [
        (ws_inv, total_invoices + 1, len(inv_headers)),
        (ws_items, total_items + 1, len(item_headers)),
    ]:
        for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
            for cell in row:
                cell.border = thin_border

    wb.save(output_path)
    print(f"\nMaintenance XLSX saved: {output_path}")
    print(f"  Sheet 'Invoices': {total_invoices} invoices (totals, vehicle, work descriptions)")
    print(f"  Sheet 'Line Items': {total_items} parts/services (linked by Invoice ID)")
    print(f"  Sum 'Invoice Total' on Invoices sheet for accurate totals")
    print(f"  Sum 'Line Amount' on Line Items sheet for part-level totals")
    return output_path


def push_to_api(pdf_path, api_url="http://localhost:3000"):
    """Extract invoice and POST to the PipelineX chief_records API
    as a maintenance_tracker collection row."""
    import urllib.request

    data = extract_invoice_data(pdf_path)
    row = to_maintenance_tracker_row(data)

    # Use the chief import save endpoint
    url = f"{api_url}/api/chief/import/save"
    payload = {
        "collection": "maintenance_tracker",
        "rows": [row],
    }

    json_data = json.dumps(payload).encode("utf-8")
    req = urllib.request.Request(url, method="POST",
                                 headers={"Content-Type": "application/json"},
                                 data=json_data)

    try:
        resp = urllib.request.urlopen(req)
        result = json.loads(resp.read().decode())
        print(f"Imported to PipelineX maintenance_tracker:")
        print(f"  Vendor: {data['vendor']}")
        print(f"  Unit: {data['unit_number']}")
        print(f"  Total: ${data.get('grand_total', 0)}")
        print(f"  Line items: {len(data['line_items'])}")
        return result
    except Exception as e:
        print(f"API import failed: {e}")
        return None


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage:")
        print("  python extract_invoice.py <invoice.pdf>                    # Excel per-invoice")
        print("  python extract_invoice.py <invoice.pdf> --json             # + JSON export")
        print("  python extract_invoice.py <invoice.pdf> --maintenance      # Maintenance tracker XLSX")
        print("  python extract_invoice.py *.pdf --maintenance              # All PDFs -> one XLSX")
        print("  python extract_invoice.py <invoice.pdf> --import           # Push to PipelineX API")
        print("  python extract_invoice.py <invoice.pdf> --import --url http://localhost:3000")
        sys.exit(0)

    args = [a for a in sys.argv[1:] if a.startswith("--")]
    pdf_paths = [a for a in sys.argv[1:] if not a.startswith("--")]

    # Maintenance tracker XLSX (for PipelineX upload)
    if "--maintenance" in args:
        export_maintenance_xlsx(pdf_paths)
    else:
        # Per-invoice Excel
        for pdf_path in pdf_paths:
            try:
                extract_invoice(pdf_path)
            except PermissionError:
                print(f"  Skipping Excel (file open): {pdf_path}")

    # JSON export
    if "--json" in args:
        for pdf_path in pdf_paths:
            export_json(pdf_path)

    # API import
    if "--import" in args:
        api_url = "http://localhost:3000"
        if "--url" in args:
            url_idx = args.index("--url")
            if url_idx + 1 < len(args):
                api_url = args[url_idx + 1]
        for pdf_path in pdf_paths:
            push_to_api(pdf_path, api_url)
                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                    