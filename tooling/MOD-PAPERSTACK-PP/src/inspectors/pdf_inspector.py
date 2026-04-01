"""
PDF Inspector - Visual PDF Reverse Engineering Tool
=====================================================
TRUE NORTH DATA STRATEGIES

WHAT THIS DOES:
  Opens a web browser with a split-screen view:
    LEFT SIDE:  Visual render of your PDF page with clickable elements
    RIGHT SIDE: The extraction code for whatever you click on

  Click any text on the PDF and the right panel shows:
    - Exact x,y coordinates
    - Font name, size, color
    - The pdfplumber code to extract that specific element
    - The code to extract everything in that region (for tables/groups)
    - Ready-to-use code for writing to Excel, CSV, or database

  This is a PARSING TOOL — designed for building data extraction pipelines.
  See a field on a PDF → click it → get the code to pull that data.

HOW TO RUN:
  python pdf_inspector.py invoice.pdf
  python pdf_inspector.py report.pdf --port 8080

  Then open http://localhost:5000 in your browser (opens automatically).

REQUIRES:
  pip install pdfplumber flask pdf2image Pillow

WHAT YOU CAN DO WITH THE OUTPUT:
  - Build PDF-to-spreadsheet parsers
  - Extract specific fields from invoices, forms, reports
  - Map PDF coordinates for automated data extraction
  - Create repeatable parsing scripts for any PDF format
"""

import sys
import os
import json
import base64
import io
import webbrowser
import threading

# ============================================
# pdfplumber: Reads PDFs and extracts every
# character, word, line, and rectangle with
# exact positions, fonts, sizes, and colors
# ============================================
import pdfplumber

# ============================================
# pdf2image: Converts PDF pages to images
# so we can display them in the browser
# ============================================
from pdf2image import convert_from_path

# ============================================
# Flask: Lightweight web server that serves
# the inspector UI to your browser
# ============================================
from flask import Flask, render_template_string, jsonify, request

# ============================================
# Pillow (PIL): Image processing - we use it
# to convert the rendered page to base64
# for embedding in the HTML
# ============================================
from PIL import Image


# ============================================
# EXTRACT ALL ELEMENTS FROM A PDF
# ============================================
def extract_pdf_data(filepath):
    """
    Reads a PDF and extracts every text element with full positioning data.
    Returns a list of pages, each containing words with coordinates.
    """
    pdf = pdfplumber.open(filepath)
    pages_data = []

    for page_num, page in enumerate(pdf.pages):
        page_info = {
            "page_num": page_num + 1,
            "width": float(page.width),
            "height": float(page.height),
            "words": [],
            "rects": [],
            "lines": [],
            "chars": [],
        }

        # Extract words (grouped characters) with positioning
        words = page.extract_words(
            keep_blank_chars=True,
            extra_attrs=["fontname", "size", "stroking_color", "non_stroking_color"]
        )
        for w in words:
            word = {
                "text": w["text"],
                "x0": round(float(w["x0"]), 2),         # Left edge
                "top": round(float(w["top"]), 2),        # Top edge (distance from page top)
                "x1": round(float(w["x1"]), 2),          # Right edge
                "bottom": round(float(w["bottom"]), 2),  # Bottom edge
                "width": round(float(w["x1"]) - float(w["x0"]), 2),
                "height": round(float(w["bottom"]) - float(w["top"]), 2),
                "fontname": w.get("fontname", "Unknown"),
                "size": round(float(w.get("size", 0)), 1),
            }

            # Calculate center point (useful for region-based extraction)
            word["cx"] = round((word["x0"] + word["x1"]) / 2, 2)
            word["cy"] = round((word["top"] + word["bottom"]) / 2, 2)

            page_info["words"].append(word)

        # Extract rectangles (table borders, shaded regions)
        if page.rects:
            for r in page.rects:
                page_info["rects"].append({
                    "x0": round(float(r["x0"]), 2),
                    "top": round(float(r["top"]), 2),
                    "x1": round(float(r["x1"]), 2),
                    "bottom": round(float(r["bottom"]), 2),
                    "fill": str(r.get("fill", "")) if r.get("fill") else None,
                    "stroke": str(r.get("stroking_color", "")) if r.get("stroking_color") else None,
                })

        # Extract lines (table gridlines, separators)
        if page.lines:
            for l in page.lines:
                page_info["lines"].append({
                    "x0": round(float(l["x0"]), 2),
                    "top": round(float(l["top"]), 2),
                    "x1": round(float(l["x1"]), 2),
                    "bottom": round(float(l["bottom"]), 2),
                })

        # Try to extract tables
        tables = page.extract_tables()
        page_info["tables"] = []
        if tables:
            for t_idx, table in enumerate(tables):
                page_info["tables"].append({
                    "index": t_idx,
                    "rows": len(table),
                    "cols": len(table[0]) if table else 0,
                    "data": table,
                })

        pages_data.append(page_info)

    pdf.close()
    return pages_data


def render_page_image(filepath, page_num, dpi=200):
    """
    Converts a PDF page to a PNG image and returns it as a base64 string.
    The base64 string can be embedded directly in HTML as an <img> src.
    """
    images = convert_from_path(filepath, dpi=dpi, first_page=page_num, last_page=page_num)
    if images:
        img = images[0]
        buffer = io.BytesIO()
        img.save(buffer, format="PNG")
        b64 = base64.b64encode(buffer.getvalue()).decode("utf-8")
        return b64, img.width, img.height
    return None, 0, 0


# ============================================
# FLASK WEB SERVER + UI
# ============================================

HTML_TEMPLATE = """
<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>PDF Inspector — {{ filename }}</title>
<style>
    /* ============================================
     * CSS VARIABLES
     * ============================================ */
    :root {
        --bg: #0d1117;
        --bg-panel: #161b22;
        --bg-code: #1c2129;
        --border: #30363d;
        --text: #e6edf3;
        --text-muted: #8b949e;
        --teal: #3d8eb9;
        --teal-glow: rgba(61, 142, 185, 0.3);
        --navy: #1a3a5c;
        --green: #3fb950;
        --yellow: #d29922;
        --red: #f85149;
        --orange: #db6d28;
    }

    * { margin: 0; padding: 0; box-sizing: border-box; }

    body {
        font-family: 'Segoe UI', -apple-system, sans-serif;
        background: var(--bg);
        color: var(--text);
        height: 100vh;
        overflow: hidden;
    }

    /* Top toolbar */
    .toolbar {
        display: flex;
        align-items: center;
        justify-content: space-between;
        padding: 8px 16px;
        background: var(--bg-panel);
        border-bottom: 1px solid var(--border);
        height: 48px;
    }

    .toolbar-left {
        display: flex;
        align-items: center;
        gap: 16px;
    }

    .toolbar h1 {
        font-size: 14px;
        font-weight: 600;
        color: var(--teal);
        letter-spacing: 1px;
    }

    .toolbar .filename {
        font-size: 13px;
        color: var(--text-muted);
    }

    .toolbar .page-nav {
        display: flex;
        align-items: center;
        gap: 8px;
    }

    .toolbar .page-nav button {
        background: var(--bg);
        color: var(--text);
        border: 1px solid var(--border);
        padding: 4px 12px;
        border-radius: 4px;
        cursor: pointer;
        font-size: 12px;
    }

    .toolbar .page-nav button:hover { border-color: var(--teal); }
    .toolbar .page-nav span { font-size: 13px; color: var(--text-muted); }

    /* Main split view */
    .split-container {
        display: flex;
        height: calc(100vh - 48px);
    }

    /* Left panel - PDF view */
    .pdf-panel {
        flex: 1;
        overflow: auto;
        position: relative;
        background: #1a1a2e;
        border-right: 1px solid var(--border);
    }

    .pdf-container {
        position: relative;
        display: inline-block;
        cursor: crosshair;
    }

    .pdf-container img {
        display: block;
    }

    /* Clickable overlay rectangles for each word */
    .word-overlay {
        position: absolute;
        border: 1px solid transparent;
        cursor: pointer;
        transition: all 0.15s;
    }

    .word-overlay:hover {
        background: rgba(61, 142, 185, 0.25);
        border-color: var(--teal);
    }

    .word-overlay.selected {
        background: rgba(61, 142, 185, 0.35);
        border: 2px solid var(--teal);
        box-shadow: 0 0 8px var(--teal-glow);
    }

    .word-overlay.region-selected {
        background: rgba(59, 185, 80, 0.2);
        border: 1px solid var(--green);
    }

    /* Coordinate tooltip that follows cursor */
    .coord-tooltip {
        position: fixed;
        background: var(--bg-panel);
        border: 1px solid var(--border);
        padding: 4px 8px;
        border-radius: 4px;
        font-size: 11px;
        font-family: 'Consolas', 'SF Mono', monospace;
        color: var(--teal);
        pointer-events: none;
        display: none;
        z-index: 100;
    }

    /* Region selection box */
    .region-select {
        position: absolute;
        border: 2px dashed var(--green);
        background: rgba(59, 185, 80, 0.1);
        pointer-events: none;
        display: none;
    }

    /* Right panel - Code view */
    .code-panel {
        width: 520px;
        min-width: 400px;
        display: flex;
        flex-direction: column;
        background: var(--bg-panel);
    }

    .code-tabs {
        display: flex;
        border-bottom: 1px solid var(--border);
    }

    .code-tab {
        padding: 8px 16px;
        font-size: 12px;
        color: var(--text-muted);
        cursor: pointer;
        border-bottom: 2px solid transparent;
        background: none;
        border-top: none;
        border-left: none;
        border-right: none;
    }

    .code-tab.active {
        color: var(--teal);
        border-bottom-color: var(--teal);
    }

    .code-tab:hover { color: var(--text); }

    .code-content {
        flex: 1;
        overflow: auto;
        padding: 16px;
    }

    /* Element info panel */
    .element-info {
        padding: 12px 16px;
        border-bottom: 1px solid var(--border);
        font-size: 12px;
        display: none;
    }

    .element-info.visible { display: block; }

    .info-grid {
        display: grid;
        grid-template-columns: repeat(3, 1fr);
        gap: 8px;
        margin-top: 8px;
    }

    .info-item {
        background: var(--bg);
        padding: 6px 10px;
        border-radius: 4px;
        border: 1px solid var(--border);
    }

    .info-item .label {
        font-size: 10px;
        color: var(--text-muted);
        text-transform: uppercase;
        letter-spacing: 0.5px;
    }

    .info-item .value {
        font-size: 13px;
        color: var(--text);
        font-family: 'Consolas', monospace;
        margin-top: 2px;
    }

    .info-item .value.highlight { color: var(--teal); }

    pre.code-block {
        background: var(--bg-code);
        border: 1px solid var(--border);
        border-radius: 6px;
        padding: 14px;
        font-size: 12.5px;
        font-family: 'Consolas', 'SF Mono', 'Fira Code', monospace;
        line-height: 1.6;
        overflow-x: auto;
        white-space: pre;
        margin-bottom: 12px;
        color: var(--text);
    }

    .code-section-title {
        font-size: 11px;
        font-weight: 600;
        color: var(--teal);
        text-transform: uppercase;
        letter-spacing: 1px;
        margin-bottom: 8px;
        margin-top: 16px;
    }

    .code-section-title:first-child { margin-top: 0; }

    .empty-state {
        display: flex;
        flex-direction: column;
        align-items: center;
        justify-content: center;
        height: 100%;
        color: var(--text-muted);
        text-align: center;
        padding: 32px;
    }

    .empty-state .icon { font-size: 48px; margin-bottom: 16px; opacity: 0.5; }
    .empty-state h3 { font-size: 16px; margin-bottom: 8px; color: var(--text); }
    .empty-state p { font-size: 13px; line-height: 1.6; }

    /* Copy button */
    .copy-btn {
        position: absolute;
        top: 8px;
        right: 8px;
        background: var(--bg-panel);
        border: 1px solid var(--border);
        color: var(--text-muted);
        padding: 4px 10px;
        border-radius: 4px;
        font-size: 11px;
        cursor: pointer;
    }
    .copy-btn:hover { color: var(--teal); border-color: var(--teal); }

    .code-wrapper { position: relative; }

    /* Stats bar at bottom */
    .stats-bar {
        padding: 6px 16px;
        background: var(--bg);
        border-top: 1px solid var(--border);
        font-size: 11px;
        color: var(--text-muted);
        display: flex;
        gap: 16px;
    }

    .stats-bar .stat-val { color: var(--teal); }

    /* Mode toggle */
    .mode-toggle {
        display: flex;
        gap: 4px;
        background: var(--bg);
        border-radius: 4px;
        padding: 2px;
    }
    .mode-btn {
        padding: 4px 12px;
        font-size: 11px;
        border: none;
        background: transparent;
        color: var(--text-muted);
        cursor: pointer;
        border-radius: 3px;
    }
    .mode-btn.active { background: var(--teal); color: white; }
</style>
</head>
<body>

<!-- Top toolbar -->
<div class="toolbar">
    <div class="toolbar-left">
        <h1>PDF INSPECTOR</h1>
        <span class="filename">{{ filename }}</span>
        <div class="page-nav">
            <button onclick="changePage(-1)">◀ Prev</button>
            <span id="pageIndicator">Page 1 / {{ total_pages }}</span>
            <button onclick="changePage(1)">Next ▶</button>
        </div>
    </div>
    <div style="display:flex; gap:12px; align-items:center;">
        <div class="mode-toggle">
            <button class="mode-btn active" onclick="setMode('click', this)">Click</button>
            <button class="mode-btn" onclick="setMode('region', this)">Region</button>
        </div>
        <span style="font-size:11px; color:var(--text-muted);">
            True North Data Strategies
        </span>
    </div>
</div>

<!-- Split view -->
<div class="split-container">

    <!-- LEFT: PDF View -->
    <div class="pdf-panel" id="pdfPanel">
        <div class="pdf-container" id="pdfContainer">
            <img id="pdfImage" src="" alt="PDF Page">
            <!-- Word overlays get injected here by JavaScript -->
            <div class="region-select" id="regionSelect"></div>
        </div>
    </div>

    <!-- RIGHT: Code View -->
    <div class="code-panel">
        <!-- Element info bar -->
        <div class="element-info" id="elementInfo">
            <div style="font-weight:600; margin-bottom:4px;">
                <span style="color:var(--teal);">■</span> Selected Element
            </div>
            <div class="info-grid" id="infoGrid"></div>
        </div>

        <!-- Code tabs -->
        <div class="code-tabs">
            <button class="code-tab active" onclick="switchTab('extract', this)">Extract Code</button>
            <button class="code-tab" onclick="switchTab('region', this)">Region Code</button>
            <button class="code-tab" onclick="switchTab('table', this)">Table Code</button>
            <button class="code-tab" onclick="switchTab('export', this)">Export Code</button>
        </div>

        <!-- Code content -->
        <div class="code-content" id="codeContent">
            <div class="empty-state">
                <div class="icon">⎔</div>
                <h3>Click any text on the PDF</h3>
                <p>
                    Click a word to see its exact coordinates and extraction code.
                    <br><br>
                    Switch to <strong>Region</strong> mode to drag-select an area
                    and get code for all elements in that zone.
                    <br><br>
                    Every code snippet is ready to copy into your parsing script.
                </p>
            </div>
        </div>

        <!-- Stats bar -->
        <div class="stats-bar">
            <span>Words: <span class="stat-val" id="statWords">0</span></span>
            <span>Tables: <span class="stat-val" id="statTables">0</span></span>
            <span>Rects: <span class="stat-val" id="statRects">0</span></span>
            <span>Cursor: <span class="stat-val" id="statCursor">—</span></span>
        </div>
    </div>
</div>

<!-- Coordinate tooltip -->
<div class="coord-tooltip" id="coordTooltip"></div>

<script>
// ============================================
// APP STATE
// ============================================
const pagesData = {{ pages_data | tojson }};
const filename = "{{ filename }}";
const totalPages = {{ total_pages }};
let currentPage = 0;
let mode = "click";           // "click" or "region"
let selectedWord = null;
let regionStart = null;
let regionWords = [];
let isDrawing = false;
const DPI = {{ dpi }};

// Scale factor: PDF coordinates (72 dpi) to rendered image pixels
const scaleFactor = DPI / 72.0;

// ============================================
// INITIALIZATION
// ============================================
function init() {
    loadPage(0);
    document.getElementById("statWords").textContent = pagesData[0].words.length;
    document.getElementById("statTables").textContent = pagesData[0].tables.length;
    document.getElementById("statRects").textContent = pagesData[0].rects.length;

    // Mouse tracking for coordinate tooltip
    const panel = document.getElementById("pdfPanel");
    panel.addEventListener("mousemove", (e) => {
        const tooltip = document.getElementById("coordTooltip");
        const container = document.getElementById("pdfContainer");
        const rect = container.getBoundingClientRect();
        const pdfX = ((e.clientX - rect.left) / scaleFactor).toFixed(1);
        const pdfY = ((e.clientY - rect.top) / scaleFactor).toFixed(1);
        tooltip.style.left = (e.clientX + 14) + "px";
        tooltip.style.top = (e.clientY - 28) + "px";
        tooltip.style.display = "block";
        tooltip.textContent = `x: ${pdfX}  y: ${pdfY}`;
        document.getElementById("statCursor").textContent = `(${pdfX}, ${pdfY})`;
    });

    panel.addEventListener("mouseleave", () => {
        document.getElementById("coordTooltip").style.display = "none";
    });

    // Region selection events
    const container = document.getElementById("pdfContainer");
    container.addEventListener("mousedown", onRegionStart);
    container.addEventListener("mousemove", onRegionDrag);
    container.addEventListener("mouseup", onRegionEnd);
}

// ============================================
// PAGE LOADING
// ============================================
function loadPage(pageIdx) {
    currentPage = pageIdx;
    const img = document.getElementById("pdfImage");
    img.src = `/page_image/${pageIdx}`;

    img.onload = () => {
        buildOverlays();
    };

    document.getElementById("pageIndicator").textContent =
        `Page ${pageIdx + 1} / ${totalPages}`;
    document.getElementById("statWords").textContent = pagesData[pageIdx].words.length;
    document.getElementById("statTables").textContent = pagesData[pageIdx].tables.length;
    document.getElementById("statRects").textContent = pagesData[pageIdx].rects.length;

    // Clear selection
    selectedWord = null;
    regionWords = [];
    document.getElementById("elementInfo").classList.remove("visible");
    showEmptyState();
}

function changePage(delta) {
    const newPage = currentPage + delta;
    if (newPage >= 0 && newPage < totalPages) {
        loadPage(newPage);
    }
}

// ============================================
// OVERLAY BUILDING
// Creates clickable rectangles over each word
// ============================================
function buildOverlays() {
    const container = document.getElementById("pdfContainer");
    // Remove old overlays
    container.querySelectorAll(".word-overlay").forEach(el => el.remove());

    const words = pagesData[currentPage].words;
    words.forEach((w, idx) => {
        const div = document.createElement("div");
        div.className = "word-overlay";
        // Convert PDF coordinates to pixel positions on the rendered image
        div.style.left = (w.x0 * scaleFactor) + "px";
        div.style.top = (w.top * scaleFactor) + "px";
        div.style.width = (w.width * scaleFactor) + "px";
        div.style.height = (w.height * scaleFactor) + "px";
        div.dataset.idx = idx;
        div.addEventListener("click", (e) => {
            e.stopPropagation();
            if (mode === "click") selectWord(idx);
        });
        container.appendChild(div);
    });
}

// ============================================
// WORD SELECTION (Click Mode)
// ============================================
function selectWord(idx) {
    // Clear previous selection
    document.querySelectorAll(".word-overlay.selected").forEach(el =>
        el.classList.remove("selected"));
    document.querySelectorAll(".word-overlay.region-selected").forEach(el =>
        el.classList.remove("region-selected"));

    // Highlight new selection
    const overlay = document.querySelector(`.word-overlay[data-idx="${idx}"]`);
    if (overlay) overlay.classList.add("selected");

    selectedWord = pagesData[currentPage].words[idx];
    regionWords = [selectedWord];

    showElementInfo(selectedWord);
    generateCode(selectedWord);
}

// ============================================
// REGION SELECTION (Drag Mode)
// ============================================
function onRegionStart(e) {
    if (mode !== "region") return;
    const container = document.getElementById("pdfContainer");
    const rect = container.getBoundingClientRect();
    regionStart = {
        x: e.clientX - rect.left,
        y: e.clientY - rect.top,
        clientX: e.clientX,
        clientY: e.clientY
    };
    isDrawing = true;

    const sel = document.getElementById("regionSelect");
    sel.style.left = regionStart.x + "px";
    sel.style.top = regionStart.y + "px";
    sel.style.width = "0px";
    sel.style.height = "0px";
    sel.style.display = "block";
}

function onRegionDrag(e) {
    if (!isDrawing || mode !== "region") return;
    const container = document.getElementById("pdfContainer");
    const rect = container.getBoundingClientRect();
    const currentX = e.clientX - rect.left;
    const currentY = e.clientY - rect.top;

    const sel = document.getElementById("regionSelect");
    const x = Math.min(regionStart.x, currentX);
    const y = Math.min(regionStart.y, currentY);
    const w = Math.abs(currentX - regionStart.x);
    const h = Math.abs(currentY - regionStart.y);

    sel.style.left = x + "px";
    sel.style.top = y + "px";
    sel.style.width = w + "px";
    sel.style.height = h + "px";
}

function onRegionEnd(e) {
    if (!isDrawing || mode !== "region") return;
    isDrawing = false;

    const container = document.getElementById("pdfContainer");
    const rect = container.getBoundingClientRect();
    const endX = e.clientX - rect.left;
    const endY = e.clientY - rect.top;

    // Convert pixel coordinates to PDF coordinates
    const pdfX0 = Math.min(regionStart.x, endX) / scaleFactor;
    const pdfY0 = Math.min(regionStart.y, endY) / scaleFactor;
    const pdfX1 = Math.max(regionStart.x, endX) / scaleFactor;
    const pdfY1 = Math.max(regionStart.y, endY) / scaleFactor;

    // Find all words in this region
    document.querySelectorAll(".word-overlay.selected, .word-overlay.region-selected")
        .forEach(el => el.classList.remove("selected", "region-selected"));

    regionWords = [];
    const words = pagesData[currentPage].words;
    words.forEach((w, idx) => {
        if (w.x0 >= pdfX0 - 2 && w.x1 <= pdfX1 + 2 &&
            w.top >= pdfY0 - 2 && w.bottom <= pdfY1 + 2) {
            regionWords.push(w);
            const overlay = document.querySelector(`.word-overlay[data-idx="${idx}"]`);
            if (overlay) overlay.classList.add("region-selected");
        }
    });

    if (regionWords.length > 0) {
        generateRegionCode(regionWords, pdfX0, pdfY0, pdfX1, pdfY1);
    }

    document.getElementById("regionSelect").style.display = "none";
}

// ============================================
// ELEMENT INFO DISPLAY
// ============================================
function showElementInfo(w) {
    const info = document.getElementById("elementInfo");
    const grid = document.getElementById("infoGrid");
    info.classList.add("visible");

    grid.innerHTML = `
        <div class="info-item">
            <div class="label">Text</div>
            <div class="value highlight">${escapeHtml(w.text.substring(0, 30))}</div>
        </div>
        <div class="info-item">
            <div class="label">Position</div>
            <div class="value">(${w.x0}, ${w.top})</div>
        </div>
        <div class="info-item">
            <div class="label">Size</div>
            <div class="value">${w.width} × ${w.height}</div>
        </div>
        <div class="info-item">
            <div class="label">Font</div>
            <div class="value">${w.fontname}</div>
        </div>
        <div class="info-item">
            <div class="label">Font Size</div>
            <div class="value">${w.size}pt</div>
        </div>
        <div class="info-item">
            <div class="label">Bounds</div>
            <div class="value">(${w.x0},${w.top})→(${w.x1},${w.bottom})</div>
        </div>
    `;
}

// ============================================
// CODE GENERATION
// ============================================
function generateCode(w) {
    const page = currentPage + 1;
    const content = document.getElementById("codeContent");

    content.innerHTML = `
        <div class="code-section-title">Extract This Exact Element</div>
        <div class="code-wrapper">
            <button class="copy-btn" onclick="copyCode(this)">Copy</button>
            <pre class="code-block">import pdfplumber

pdf = pdfplumber.open("${filename}")
page = pdf.pages[${currentPage}]  # Page ${page}

# Crop to exact region of this element
# Coordinates: x0=${w.x0}, top=${w.top}, x1=${w.x1}, bottom=${w.bottom}
region = page.crop((${w.x0}, ${w.top}, ${w.x1}, ${w.bottom}))
text = region.extract_text()
print(text)  # → "${escapeHtml(w.text)}"

pdf.close()</pre>
        </div>

        <div class="code-section-title">Find by Position (Reusable for Same PDF Layout)</div>
        <div class="code-wrapper">
            <button class="copy-btn" onclick="copyCode(this)">Copy</button>
            <pre class="code-block">import pdfplumber

def extract_field(pdf_path, page_num, x0, top, x1, bottom):
    # Extract text from a specific region of a PDF page.
    pdf = pdfplumber.open(pdf_path)
    page = pdf.pages[page_num]
    region = page.crop((x0, top, x1, bottom))
    text = region.extract_text()
    pdf.close()
    return text

# Usage — coordinates from "${escapeHtml(w.text)}"
value = extract_field(
    "${filename}",
    page_num=${currentPage},
    x0=${w.x0}, top=${w.top},
    x1=${w.x1}, bottom=${w.bottom}
)
print(f"Extracted: {value}")</pre>
        </div>

        <div class="code-section-title">Find by Font/Size (All Text Matching This Style)</div>
        <div class="code-wrapper">
            <button class="copy-btn" onclick="copyCode(this)">Copy</button>
            <pre class="code-block">import pdfplumber

pdf = pdfplumber.open("${filename}")
page = pdf.pages[${currentPage}]

# Find all words with same font and size
# Font: ${w.fontname}, Size: ${w.size}pt
words = page.extract_words(extra_attrs=["fontname", "size"])
matches = [
    w for w in words
    if w["fontname"] == "${w.fontname}" and round(w["size"], 1) == ${w.size}
]

for m in matches:
    print(f"  '{m['text']}'  at ({m['x0']:.1f}, {m['top']:.1f})")

pdf.close()</pre>
        </div>

        <div class="code-section-title">Write to Excel</div>
        <div class="code-wrapper">
            <button class="copy-btn" onclick="copyCode(this)">Copy</button>
            <pre class="code-block">import pdfplumber
import openpyxl  # pip install openpyxl

pdf = pdfplumber.open("${filename}")
page = pdf.pages[${currentPage}]

# Extract this field
region = page.crop((${w.x0}, ${w.top}, ${w.x1}, ${w.bottom}))
value = region.extract_text()

# Write to Excel
wb = openpyxl.Workbook()
ws = wb.active
ws["A1"] = "Field"
ws["B1"] = "Value"
ws["A2"] = "${escapeHtml(w.text.substring(0, 30))}"
ws["B2"] = value
wb.save("extracted_data.xlsx")
print("Saved to extracted_data.xlsx")

pdf.close()</pre>
        </div>
    `;
}

function generateRegionCode(words, x0, y0, x1, y1) {
    const page = currentPage + 1;
    const content = document.getElementById("codeContent");

    // Show info
    const info = document.getElementById("elementInfo");
    const grid = document.getElementById("infoGrid");
    info.classList.add("visible");
    grid.innerHTML = `
        <div class="info-item">
            <div class="label">Region</div>
            <div class="value highlight">${words.length} words selected</div>
        </div>
        <div class="info-item">
            <div class="label">Bounds</div>
            <div class="value">(${x0.toFixed(1)},${y0.toFixed(1)})→(${x1.toFixed(1)},${y1.toFixed(1)})</div>
        </div>
        <div class="info-item">
            <div class="label">Size</div>
            <div class="value">${(x1-x0).toFixed(1)} × ${(y1-y0).toFixed(1)} pts</div>
        </div>
    `;

    // Build word list preview
    const wordPreview = words.slice(0, 10).map(w => `  "${w.text}"`).join("\\n");
    const fullText = words.map(w => w.text).join(" ");

    content.innerHTML = `
        <div class="code-section-title">Extract This Region</div>
        <div class="code-wrapper">
            <button class="copy-btn" onclick="copyCode(this)">Copy</button>
            <pre class="code-block">import pdfplumber

pdf = pdfplumber.open("${filename}")
page = pdf.pages[${currentPage}]  # Page ${page}

# Region: (${x0.toFixed(1)}, ${y0.toFixed(1)}) to (${x1.toFixed(1)}, ${y1.toFixed(1)})
# Contains ${words.length} words
region = page.crop((${x0.toFixed(1)}, ${y0.toFixed(1)}, ${x1.toFixed(1)}, ${y1.toFixed(1)}))
text = region.extract_text()
print(text)

pdf.close()</pre>
        </div>

        <div class="code-section-title">Extract as Table (If Region Contains a Table)</div>
        <div class="code-wrapper">
            <button class="copy-btn" onclick="copyCode(this)">Copy</button>
            <pre class="code-block">import pdfplumber

pdf = pdfplumber.open("${filename}")
page = pdf.pages[${currentPage}]

region = page.crop((${x0.toFixed(1)}, ${y0.toFixed(1)}, ${x1.toFixed(1)}, ${y1.toFixed(1)}))
table = region.extract_table()

if table:
    for row in table:
        print(row)
else:
    print("No table structure detected in region")
    print("Plain text:", region.extract_text())

pdf.close()</pre>
        </div>

        <div class="code-section-title">Region to Excel (All Words with Positions)</div>
        <div class="code-wrapper">
            <button class="copy-btn" onclick="copyCode(this)">Copy</button>
            <pre class="code-block">import pdfplumber
import openpyxl  # pip install openpyxl

pdf = pdfplumber.open("${filename}")
page = pdf.pages[${currentPage}]

region = page.crop((${x0.toFixed(1)}, ${y0.toFixed(1)}, ${x1.toFixed(1)}, ${y1.toFixed(1)}))
words = region.extract_words(extra_attrs=["fontname", "size"])

wb = openpyxl.Workbook()
ws = wb.active
ws.append(["Text", "X0", "Top", "X1", "Bottom", "Font", "Size"])

for w in words:
    ws.append([w["text"], w["x0"], w["top"], w["x1"], w["bottom"],
               w.get("fontname", ""), w.get("size", "")])

wb.save("region_extract.xlsx")
print(f"Exported {len(words)} words to region_extract.xlsx")

pdf.close()</pre>
        </div>

        <div class="code-section-title">Region to CSV</div>
        <div class="code-wrapper">
            <button class="copy-btn" onclick="copyCode(this)">Copy</button>
            <pre class="code-block">import pdfplumber
import csv

pdf = pdfplumber.open("${filename}")
page = pdf.pages[${currentPage}]

region = page.crop((${x0.toFixed(1)}, ${y0.toFixed(1)}, ${x1.toFixed(1)}, ${y1.toFixed(1)}))

# Option A: Extract as table
table = region.extract_table()
if table:
    with open("region_data.csv", "w", newline="") as f:
        writer = csv.writer(f)
        for row in table:
            writer.writerow(row)
    print("Table saved to region_data.csv")

# Option B: Extract as positioned words
words = region.extract_words()
with open("region_words.csv", "w", newline="") as f:
    writer = csv.writer(f)
    writer.writerow(["text", "x0", "top", "x1", "bottom"])
    for w in words:
        writer.writerow([w["text"], w["x0"], w["top"], w["x1"], w["bottom"]])
print("Words saved to region_words.csv")

pdf.close()</pre>
        </div>

        <div class="code-section-title">Words Found in Region</div>
        <pre class="code-block">${words.map(w =>
            `"${escapeHtml(w.text)}"  x0=${w.x0} top=${w.top} font=${w.fontname} size=${w.size}pt`
        ).join("\\n")}</pre>
    `;
}

function showEmptyState() {
    document.getElementById("codeContent").innerHTML = `
        <div class="empty-state">
            <div class="icon">⎔</div>
            <h3>Click any text on the PDF</h3>
            <p>
                Click a word to see its exact coordinates and extraction code.
                <br><br>
                Switch to <strong>Region</strong> mode to drag-select an area
                and get code for all elements in that zone.
            </p>
        </div>
    `;
}

// ============================================
// UI HELPERS
// ============================================
function setMode(newMode, btn) {
    mode = newMode;
    document.querySelectorAll(".mode-btn").forEach(b => b.classList.remove("active"));
    btn.classList.add("active");
    document.getElementById("pdfContainer").style.cursor =
        mode === "region" ? "crosshair" : "crosshair";
}

function switchTab(tab, btn) {
    document.querySelectorAll(".code-tab").forEach(t => t.classList.remove("active"));
    btn.classList.add("active");

    // Re-trigger code generation for the selected tab
    if (selectedWord && tab === "extract") generateCode(selectedWord);
    if (regionWords.length > 0 && tab === "region") {
        const ws = regionWords;
        const x0 = Math.min(...ws.map(w => w.x0));
        const y0 = Math.min(...ws.map(w => w.top));
        const x1 = Math.max(...ws.map(w => w.x1));
        const y1 = Math.max(...ws.map(w => w.bottom));
        generateRegionCode(ws, x0, y0, x1, y1);
    }
}

function copyCode(btn) {
    const pre = btn.nextElementSibling;
    navigator.clipboard.writeText(pre.textContent).then(() => {
        btn.textContent = "Copied!";
        setTimeout(() => btn.textContent = "Copy", 1500);
    });
}

function escapeHtml(text) {
    return text.replace(/&/g, "&amp;").replace(/</g, "&lt;")
               .replace(/>/g, "&gt;").replace(/"/g, "&quot;");
}

// Boot
init();
</script>
</body>
</html>
"""

# ============================================
# FLASK APP
# ============================================
def create_app(filepath):
    """Create and configure the Flask web application."""
    app = Flask(__name__)
    app.config["FILEPATH"] = filepath

    filename = os.path.basename(filepath)
    dpi = 200                                          # Render resolution for PDF pages

    # Extract all PDF data upfront
    print(f"Extracting PDF data from: {filename}")
    pages_data = extract_pdf_data(filepath)
    total_pages = len(pages_data)
    print(f"  Pages: {total_pages}")
    for i, pd in enumerate(pages_data):
        print(f"  Page {i+1}: {len(pd['words'])} words, {len(pd['tables'])} tables, "
              f"{len(pd['rects'])} rects, {pd['width']}x{pd['height']} pts")

    # Pre-render page images
    print("Rendering page images...")
    page_images = {}
    for i in range(total_pages):
        b64, w, h = render_page_image(filepath, i + 1, dpi=dpi)
        page_images[i] = b64
        print(f"  Page {i+1}: {w}x{h} px")

    @app.route("/")
    def index():
        return render_template_string(
            HTML_TEMPLATE,
            filename=filename,
            total_pages=total_pages,
            pages_data=pages_data,
            dpi=dpi,
        )

    @app.route("/page_image/<int:page_idx>")
    def page_image(page_idx):
        from flask import Response
        if page_idx in page_images:
            img_bytes = base64.b64decode(page_images[page_idx])
            return Response(img_bytes, mimetype="image/png")
        return "Not found", 404

    return app


# ============================================
# MAIN
# ============================================
def main():
    if len(sys.argv) < 2:
        print("=" * 60)
        print("  PDF Inspector - Visual PDF Reverse Engineering")
        print("  True North Data Strategies")
        print("=" * 60)
        print()
        print("  Usage:")
        print("    python pdf_inspector.py <file.pdf> [--port 5000]")
        print()
        print("  Examples:")
        print("    python pdf_inspector.py invoice.pdf")
        print("    python pdf_inspector.py report.pdf --port 8080")
        print()
        print("  Then open http://localhost:5000 in your browser.")
        print()
        print("  Features:")
        print("    - Click any text to see coordinates + extraction code")
        print("    - Drag to select regions and get bulk extraction code")
        print("    - Code output for: pdfplumber, Excel, CSV, database")
        print("    - Live coordinate tracking as you move your mouse")
        print()
        sys.exit(0)

    filepath = sys.argv[1]
    if not os.path.exists(filepath):
        print(f"Error: File not found: {filepath}")
        sys.exit(1)

    port = 5000
    if "--port" in sys.argv:
        idx = sys.argv.index("--port")
        if idx + 1 < len(sys.argv):
            port = int(sys.argv[idx + 1])

    app = create_app(filepath)

    # Auto-open browser after slight delay (gives server time to start)
    def open_browser():
        import time
        time.sleep(1.5)
        webbrowser.open(f"http://localhost:{port}")

    threading.Thread(target=open_browser, daemon=True).start()

    print(f"\n  PDF Inspector running at: http://localhost:{port}")
    print(f"  Press Ctrl+C to stop.\n")
    app.run(host="0.0.0.0", port=port, debug=False)


if __name__ == "__main__":
    main()
