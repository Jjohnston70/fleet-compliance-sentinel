"""
Public API for the Invoice Extraction Module.

Provides clean interfaces for single and batch extraction,
and output formatting to Excel and JSON.
"""

import sys
import os
import json

# Add src/ to path
_src_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if _src_dir not in sys.path:
    sys.path.insert(0, _src_dir)

from vendor_parsers import parse_invoice


def extract(pdf_path, org_id=None, operator="system"):
    """
    Extract structured data from a single invoice PDF.
    
    Returns standardized dict with all invoice fields extracted
    by the vendor-specific parser.
    
    Args:
        pdf_path (str): Path to invoice PDF file
        org_id (str, optional): Organization ID for audit logging
        operator (str, optional): Operator identifier for audit logging
    
    Returns:
        dict: Standardized invoice data including vendor, amounts, line items, etc.
    
    Raises:
        ValueError: If PDF cannot be parsed (unknown vendor or parse error)
        FileNotFoundError: If PDF file does not exist
    """
    if not os.path.exists(pdf_path):
        raise FileNotFoundError(f"Invoice PDF not found: {pdf_path}")
    
    try:
        result = parse_invoice(pdf_path)
        if org_id:
            result["org_id"] = org_id
        if operator:
            result["operator"] = operator
        return result
    except Exception as e:
        raise ValueError(f"Failed to extract from {pdf_path}: {e}") from e


def extract_batch(pdf_paths, org_id=None, operator="system"):
    """
    Extract from multiple PDFs.
    
    Returns list of result dicts, one per successfully parsed invoice.
    Errors are collected and reported but don't stop batch processing.
    
    Args:
        pdf_paths (list): List of PDF file paths
        org_id (str, optional): Organization ID for audit logging
        operator (str, optional): Operator identifier for audit logging
    
    Returns:
        list: List of dicts with extracted invoice data
    
    Raises:
        ValueError: If no PDFs could be successfully parsed
    """
    results = []
    errors = []
    
    for pdf_path in pdf_paths:
        try:
            result = extract(pdf_path, org_id=org_id, operator=operator)
            results.append(result)
        except (FileNotFoundError, ValueError) as e:
            errors.append({"pdf_path": pdf_path, "error": str(e)})
    
    if not results and errors:
        error_summary = "; ".join(f"{e['pdf_path']}: {e['error']}" for e in errors)
        raise ValueError(f"No PDFs parsed successfully: {error_summary}")
    
    return results


def to_xlsx(data_list, output_path):
    """
    Export extracted data to Excel with Invoices + Line Items sheets.
    
    Args:
        data_list (list): List of invoice data dicts (from extract or extract_batch)
        output_path (str): Path where Excel file will be saved
    
    Returns:
        str: Path to the created Excel file
    """
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    
    if isinstance(data_list, dict):
        data_list = [data_list]
    
    if not data_list:
        raise ValueError("No data to export")
    
    wb = openpyxl.Workbook()
    
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="1A3A5C", end_color="1A3A5C", fill_type="solid")
    money_format = '#,##0.00'
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    def style_headers(ws, headers):
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
    
    # Sheet 1: Invoices
    ws_inv = wb.active
    ws_inv.title = "Invoices"
    inv_headers = [
        "Invoice ID", "Completed Date", "Vendor", "PO Number",
        "Unit Number", "Asset Name", "Year", "Make", "Model",
        "Plate Number", "VIN", "Engine", "Mileage Hours", "Written By",
        "Parts Cost", "Labor Cost", "Shop Supplies", "Sales Tax", "Invoice Total",
        "Line Item Count",
        "Work Requested", "Work Completed", "Source File",
    ]
    ws_inv.append(inv_headers)
    style_headers(ws_inv, inv_headers)
    
    # Sheet 2: Line Items
    ws_items = wb.create_sheet("Line Items")
    item_headers = [
        "Line Item ID", "Invoice ID", "Invoice Date", "Vendor",
        "Unit Number", "Asset Name", "VIN",
        "Item Qty", "Item Description", "Unit Price", "Line Amount",
    ]
    ws_items.append(item_headers)
    style_headers(ws_items, item_headers)
    
    def _generate_invoice_id(data):
        vendor = data.get("vendor", "Unknown")[:3].upper()
        date_str = data.get("invoice_date", "0000-00-00").replace("-", "")
        return f"{vendor}-{date_str}"
    
    total_invoices = 0
    total_items = 0
    
    for data in data_list:
        if not isinstance(data, dict):
            continue
        
        invoice_id = _generate_invoice_id(data)
        line_items = data.get("line_items", [])
        work_sections = data.get("work_descriptions", [])
        asset_name = f"{data.get('year', '')} {data.get('make', '')} {data.get('model', '')}".strip()
        
        work_requested = data.get("work_descriptions", [{}])[0].get("work_text", "") if work_sections else ""
        work_completed = ""
        
        ws_inv.append([
            invoice_id,
            data.get("invoice_date", ""),
            data.get("vendor", ""),
            data.get("po_number", ""),
            data.get("unit_number", ""),
            asset_name,
            data.get("year", ""),
            data.get("make", ""),
            data.get("model", ""),
            data.get("plate_number", ""),
            data.get("vin", ""),
            data.get("engine", ""),
            data.get("mileage_hours", ""),
            data.get("written_by", ""),
            data.get("parts_total", 0),
            data.get("labor_total", 0),
            data.get("shop_supplies", 0),
            data.get("sales_tax", 0),
            data.get("grand_total", 0),
            len(line_items),
            work_requested,
            work_completed,
            data.get("source_file", ""),
        ])
        total_invoices += 1
        
        for i, item in enumerate(line_items):
            line_id = f"{invoice_id}-{str(i+1).zfill(3)}"
            ws_items.append([
                line_id,
                invoice_id,
                data.get("invoice_date", ""),
                data.get("vendor", ""),
                data.get("unit_number", ""),
                asset_name,
                data.get("vin", ""),
                item.get("qty", 0),
                item.get("description", ""),
                item.get("unit_price", 0),
                item.get("amount", 0),
            ])
            total_items += 1
    
    # Format columns
    inv_money = ["Parts Cost", "Labor Cost", "Shop Supplies", "Sales Tax", "Invoice Total"]
    for col_idx, header in enumerate(inv_headers, 1):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        ws_inv.column_dimensions[col_letter].width = max(len(header) + 2, 14)
        if header in inv_money:
            for row_idx in range(2, total_invoices + 2):
                cell = ws_inv.cell(row=row_idx, column=col_idx)
                if isinstance(cell.value, (int, float)):
                    cell.number_format = money_format
    
    for header in ["Work Requested", "Work Completed", "Asset Name"]:
        if header in inv_headers:
            col_letter = openpyxl.utils.get_column_letter(inv_headers.index(header) + 1)
            ws_inv.column_dimensions[col_letter].width = 60
    
    item_money = ["Unit Price", "Line Amount"]
    for col_idx, header in enumerate(item_headers, 1):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        ws_items.column_dimensions[col_letter].width = max(len(header) + 2, 14)
        if header in item_money:
            for row_idx in range(2, total_items + 2):
                cell = ws_items.cell(row=row_idx, column=col_idx)
                if isinstance(cell.value, (int, float)):
                    cell.number_format = money_format
    
    if "Item Description" in item_headers:
        col_letter = openpyxl.utils.get_column_letter(item_headers.index("Item Description") + 1)
        ws_items.column_dimensions[col_letter].width = 55
    
    for ws, max_row, max_col in [
        (ws_inv, total_invoices + 1, len(inv_headers)),
        (ws_items, total_items + 1, len(item_headers)),
    ]:
        for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
            for cell in row:
                cell.border = thin_border
    
    wb.save(output_path)
    return output_path


def _derive_service_type(data):
    """Derive a service type label from work descriptions and vendor context."""
    descs = data.get("work_descriptions", [])
    if descs:
        text = " ".join(d.get("work_text", "") for d in descs).lower()
        if any(w in text for w in ["dot", "inspection", "annual"]):
            return "DOT Inspection"
        if any(w in text for w in ["oil", "lube", "filter", "grease"]):
            return "Lube / PM"
        if any(w in text for w in ["tire", "tyre", "mount", "balance", "flat"]):
            return "Tires"
        if any(w in text for w in ["tow", "recovery", "haul"]):
            return "Towing / Recovery"
        if any(w in text for w in ["glass", "windshield", "window"]):
            return "Glass Repair"
        if any(w in text for w in ["trans", "transmission"]):
            return "Transmission"
        if any(w in text for w in ["brake", "drum", "rotor", "pad"]):
            return "Brakes"
        if any(w in text for w in ["engine", "motor", "head gasket"]):
            return "Engine Repair"
        if any(w in text for w in ["electrical", "battery", "alternator", "starter"]):
            return "Electrical"
        if any(w in text for w in ["a/c", "hvac", "heater", "cooling"]):
            return "HVAC / Cooling"
    # Fallback: derive from vendor name
    vendor = data.get("vendor", "").lower()
    if "tire" in vendor:
        return "Tires"
    if "towing" in vendor:
        return "Towing / Recovery"
    if "glass" in vendor:
        return "Glass Repair"
    if "napa" in vendor:
        return "Parts Purchase"
    return "General Repair"


def _derive_category(data):
    """Derive invoice category from service type and amounts."""
    stype = _derive_service_type(data)
    category_map = {
        "DOT Inspection": "Compliance",
        "Lube / PM": "Preventive Maintenance",
        "Tires": "Tires & Wheels",
        "Towing / Recovery": "Towing",
        "Glass Repair": "Body & Glass",
        "Transmission": "Drivetrain",
        "Brakes": "Brakes & Suspension",
        "Engine Repair": "Engine",
        "Electrical": "Electrical",
        "HVAC / Cooling": "HVAC",
        "Parts Purchase": "Parts",
        "General Repair": "Repair",
    }
    return category_map.get(stype, "Repair")


def to_fleet_xlsx(data_list, output_path):
    """
    Export extracted data to fleet-compliance-bulk-upload-template format.

    Produces three sheets:
    - Invoices (15 columns matching fleet-compliance template)
    - Maintenance Tracker (13 columns matching fleet-compliance template)
    - Line Items (9 columns - new sheet for itemized parts/labor)

    Args:
        data_list (list): List of invoice data dicts (from extract or extract_batch)
        output_path (str): Path where Excel file will be saved

    Returns:
        str: Path to the created Excel file
    """
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    if isinstance(data_list, dict):
        data_list = [data_list]

    if not data_list:
        raise ValueError("No data to export")

    wb = openpyxl.Workbook()

    # TNDS brand styling
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="1A3A5C", end_color="1A3A5C", fill_type="solid")
    accent_fill = PatternFill(start_color="3D8EB9", end_color="3D8EB9", fill_type="solid")
    money_format = '#,##0.00'
    date_format = 'YYYY-MM-DD'
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    def style_headers(ws, headers, fill=None):
        use_fill = fill or header_fill
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = use_fill
            cell.alignment = Alignment(horizontal='center', wrap_text=True)

    def auto_width(ws, headers, min_width=12):
        for col_idx, h in enumerate(headers, 1):
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = max(len(h) + 4, min_width)

    def apply_borders(ws, max_row, max_col):
        for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
            for cell in row:
                cell.border = thin_border

    # -------------------------------------------------------
    # Sheet 1: Invoices (fleet-compliance template schema)
    # -------------------------------------------------------
    ws_inv = wb.active
    ws_inv.title = "Invoices"
    inv_headers = [
        "Vendor", "Invoice Number", "Invoice Date", "Due Date",
        "Total Amount", "Parts Cost", "Labor Cost", "Shop Supplies",
        "Sales Tax", "Category", "Asset ID", "Service Type",
        "Status", "PO Number", "Notes",
    ]
    ws_inv.append(inv_headers)
    style_headers(ws_inv, inv_headers)

    # -------------------------------------------------------
    # Sheet 2: Maintenance Tracker (fleet-compliance schema)
    # -------------------------------------------------------
    ws_maint = wb.create_sheet("Maintenance Tracker")
    maint_headers = [
        "Maintenance ID", "Asset ID", "Asset Name", "Service Type",
        "Scheduled Date", "Completed Date", "Status",
        "Parts Cost", "Labor Cost", "Total Cost",
        "Vendor", "Invoice Number", "Notes",
    ]
    ws_maint.append(maint_headers)
    style_headers(ws_maint, maint_headers, fill=accent_fill)

    # -------------------------------------------------------
    # Sheet 3: Line Items (new sheet)
    # -------------------------------------------------------
    ws_items = wb.create_sheet("Line Items")
    item_headers = [
        "Line Item ID", "Invoice Number", "Vendor", "Asset ID",
        "Qty", "Part Number", "Description", "Unit Price", "Line Amount",
    ]
    ws_items.append(item_headers)
    style_headers(ws_items, item_headers)

    total_invoices = 0
    total_maint = 0
    total_items = 0

    for idx, data in enumerate(data_list):
        if not isinstance(data, dict):
            continue

        vendor = data.get("vendor", "Unknown")
        inv_num = data.get("invoice_number", "")
        inv_date = data.get("invoice_date", "")
        unit_number = data.get("unit_number", "")
        asset_name = f"{data.get('year', '')} {data.get('make', '')} {data.get('model', '')}".strip()
        service_type = _derive_service_type(data)
        category = _derive_category(data)

        # Build notes: source file + asset identifiers
        notes_parts = []
        if data.get("source_file"):
            notes_parts.append(f"Source: {data['source_file']}")
        if data.get("vin"):
            notes_parts.append(f"VIN: {data['vin']}")
        if data.get("plate_number"):
            notes_parts.append(f"Plate: {data['plate_number']}")
        if data.get("mileage_hours"):
            notes_parts.append(f"Miles/Hrs: {data['mileage_hours']}")
        notes = " | ".join(notes_parts)

        # Work description text for maintenance tracker notes
        work_notes = ""
        work_descs = data.get("work_descriptions", [])
        if work_descs:
            work_notes = "; ".join(d.get("work_text", "") for d in work_descs if d.get("work_text"))

        # Generate maintenance ID: VND-YYYYMMDD-NNN
        vendor_prefix = vendor[:3].upper()
        date_compact = inv_date.replace("-", "").replace("/", "")[:8] if inv_date else "00000000"
        maint_id = f"{vendor_prefix}-{date_compact}-{str(idx + 1).zfill(3)}"

        # --- Invoices row ---
        ws_inv.append([
            vendor,
            inv_num,
            inv_date,
            "",  # Due Date - not extracted
            data.get("grand_total", 0),
            data.get("parts_total", 0),
            data.get("labor_total", 0),
            data.get("shop_supplies", 0),
            data.get("sales_tax", 0),
            category,
            unit_number,
            service_type,
            "Completed",
            data.get("po_number", ""),
            notes,
        ])
        total_invoices += 1

        # --- Maintenance Tracker row ---
        ws_maint.append([
            maint_id,
            unit_number,
            asset_name,
            service_type,
            "",  # Scheduled Date - not extracted
            inv_date,
            "Completed",
            data.get("parts_total", 0),
            data.get("labor_total", 0),
            data.get("grand_total", 0),
            vendor,
            inv_num,
            work_notes or notes,
        ])
        total_maint += 1

        # --- Line Items rows ---
        line_items = data.get("line_items", [])
        for i, item in enumerate(line_items):
            line_id_ref = inv_num if inv_num else maint_id
            line_id = f"{line_id_ref}-{str(i + 1).zfill(3)}"
            ws_items.append([
                line_id,
                inv_num,
                vendor,
                unit_number,
                item.get("qty", 0),
                item.get("part_number", ""),
                item.get("description", ""),
                item.get("unit_price", 0),
                item.get("amount", 0),
            ])
            total_items += 1

    # --- Formatting ---

    # Invoices: money columns
    inv_money_cols = ["Total Amount", "Parts Cost", "Labor Cost", "Shop Supplies", "Sales Tax"]
    for col_idx, header in enumerate(inv_headers, 1):
        if header in inv_money_cols:
            for row_idx in range(2, total_invoices + 2):
                cell = ws_inv.cell(row=row_idx, column=col_idx)
                if isinstance(cell.value, (int, float)):
                    cell.number_format = money_format
    auto_width(ws_inv, inv_headers)
    # Widen Notes column
    notes_col = openpyxl.utils.get_column_letter(inv_headers.index("Notes") + 1)
    ws_inv.column_dimensions[notes_col].width = 60

    # Maintenance Tracker: money columns
    maint_money_cols = ["Parts Cost", "Labor Cost", "Total Cost"]
    for col_idx, header in enumerate(maint_headers, 1):
        if header in maint_money_cols:
            for row_idx in range(2, total_maint + 2):
                cell = ws_maint.cell(row=row_idx, column=col_idx)
                if isinstance(cell.value, (int, float)):
                    cell.number_format = money_format
    auto_width(ws_maint, maint_headers)
    notes_col_m = openpyxl.utils.get_column_letter(maint_headers.index("Notes") + 1)
    ws_maint.column_dimensions[notes_col_m].width = 60

    # Line Items: money columns
    item_money_cols = ["Unit Price", "Line Amount"]
    for col_idx, header in enumerate(item_headers, 1):
        if header in item_money_cols:
            for row_idx in range(2, total_items + 2):
                cell = ws_items.cell(row=row_idx, column=col_idx)
                if isinstance(cell.value, (int, float)):
                    cell.number_format = money_format
    auto_width(ws_items, item_headers)
    desc_col = openpyxl.utils.get_column_letter(item_headers.index("Description") + 1)
    ws_items.column_dimensions[desc_col].width = 55

    # Borders on all sheets
    apply_borders(ws_inv, total_invoices + 1, len(inv_headers))
    apply_borders(ws_maint, total_maint + 1, len(maint_headers))
    apply_borders(ws_items, total_items + 1, len(item_headers))

    # Freeze top row on all sheets
    ws_inv.freeze_panes = "A2"
    ws_maint.freeze_panes = "A2"
    ws_items.freeze_panes = "A2"

    wb.save(output_path)
    return output_path


def to_json(data, indent=2):
    """
    Serialize extracted data to JSON string.
    
    Args:
        data: Single dict or list of dicts with extracted invoice data
        indent (int, optional): JSON indentation level (default: 2)
    
    Returns:
        str: JSON string representation of the data
    """
    if isinstance(data, dict):
        data = [data]
    
    return json.dumps(data, indent=indent, default=str)
