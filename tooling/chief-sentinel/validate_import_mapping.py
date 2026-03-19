import csv
from pathlib import Path

from openpyxl import load_workbook

from import_mapping import DATASET_MAPPINGS


ROOT = Path(__file__).resolve().parent


def read_csv_headers(path: Path) -> list[str]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.reader(handle)
        try:
            return next(reader)
        except StopIteration:
            return []


def read_sheet_headers(path: Path, worksheet: str) -> list[str]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook[worksheet]
        row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), [])
        return [value for value in row if value is not None]
    finally:
        workbook.close()


def validate_dataset(dataset) -> tuple[bool, list[str]]:
    errors: list[str] = []
    source_path = ROOT / dataset.source_path
    if not source_path.exists():
        return False, [f"{dataset.key}: source path not found -> {source_path}"]

    if dataset.kind == "csv":
        actual_headers = read_csv_headers(source_path)
    else:
        actual_headers = read_sheet_headers(source_path, dataset.worksheet or "")

    mapped_columns = [mapping.source_column for mapping in dataset.column_mappings]
    missing_in_mapping = [header for header in actual_headers if header not in mapped_columns]
    unknown_in_mapping = [header for header in mapped_columns if header not in actual_headers]

    if missing_in_mapping:
        errors.append(f"{dataset.key}: unmapped source headers -> {missing_in_mapping}")
    if unknown_in_mapping:
        errors.append(f"{dataset.key}: mapping references unknown headers -> {unknown_in_mapping}")

    known_headers = set(actual_headers)
    for derived in dataset.derived_mappings:
        missing_inputs = [header for header in derived.source_columns if header not in known_headers]
        if missing_inputs:
            errors.append(
                f"{dataset.key}: derived mapping {derived.collection}.{derived.field} references missing headers -> {missing_inputs}"
            )

    if dataset.kind == "config_sheet" and not dataset.config_entry_mappings:
        errors.append(f"{dataset.key}: config sheet is missing config entry mappings")

    return not errors, errors


def main() -> int:
    failures: list[str] = []
    passed = 0

    for dataset in DATASET_MAPPINGS:
        ok, errors = validate_dataset(dataset)
        if ok:
            passed += 1
            print(f"[OK] {dataset.key}")
        else:
            print(f"[FAIL] {dataset.key}")
            failures.extend(errors)

    print()
    print(f"Validated {passed}/{len(DATASET_MAPPINGS)} dataset mappings.")
    if failures:
        print()
        for failure in failures:
            print(failure)
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
