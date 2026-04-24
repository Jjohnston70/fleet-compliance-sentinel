import csv
import hashlib
import json
import re
from collections import defaultdict
from datetime import date, datetime, time
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook

from import_mapping import DATASET_MAPPINGS


ROOT = Path(__file__).resolve().parent
WEBSITE_ROOT = ROOT.parent.parent
GENERATED_DIR = ROOT / "generated_imports"
GENERATED_TS_PATH = WEBSITE_ROOT / "src" / "lib" / "chief-imported-data.generated.ts"
GENERATED_TEMPLATE_MANIFEST_PATH = GENERATED_DIR / "upload_template_manifest.json"
GENERATED_VALIDATION_REPORT_PATH = GENERATED_DIR / "validation_report.json"
GENERATED_VALIDATION_REPORT_MD_PATH = GENERATED_DIR / "validation_report.md"
GENERATED_TEMPLATE_TS_PATH = WEBSITE_ROOT / "src" / "lib" / "chief-upload-template.generated.ts"

ORG_ID = "sample-fleet"
TODAY = date(2026, 3, 17)


def normalize_scalar(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, time):
        return value.strftime("%H:%M")
    if isinstance(value, float):
        if value.is_integer():
            return int(value)
        return round(value, 2)
    if isinstance(value, str):
        text = value.strip()
        return text or None
    return value


def slugify(*values: Any) -> str | None:
    joined = " ".join(str(value).strip() for value in values if normalize_scalar(value) not in (None, ""))
    if not joined:
        return None
    slug = re.sub(r"[^a-z0-9]+", "-", joined.lower()).strip("-")
    return slug or None


def stable_hash(*values: Any) -> str:
    payload = json.dumps(values, sort_keys=True, default=str)
    return hashlib.sha1(payload.encode("utf-8")).hexdigest()[:10]


def stable_row_id(dataset_key: str, collection: str, row: dict[str, Any]) -> str:
    normalized = {
        key: normalize_scalar(value)
        for key, value in sorted(row.items())
        if normalize_scalar(value) not in (None, "", [], {})
    }
    return f"{collection.rstrip('s')}_{stable_hash(dataset_key, collection, normalized)}"


def source_row_preview(row: dict[str, Any], limit: int = 4) -> dict[str, Any]:
    preview: dict[str, Any] = {}
    for key, value in row.items():
        normalized = normalize_scalar(value)
        if normalized in (None, "", [], {}):
            continue
        preview[key] = normalized
        if len(preview) >= limit:
            break
    return preview


def make_dataset_report(dataset) -> dict[str, Any]:
    return {
        "datasetKey": dataset.key,
        "kind": dataset.kind,
        "sourcePath": dataset.source_path,
        "worksheet": dataset.worksheet,
        "rowCount": 0,
        "nonEmptyRows": 0,
        "importedRows": 0,
        "collectionWrites": defaultdict(int),
        "skippedEmptyRows": 0,
        "skippedRows": 0,
        "generatedIds": [],
        "warnings": [],
        "fallbacksUsed": [],
    }


def add_dataset_warning(
    dataset_report: dict[str, Any],
    *,
    code: str,
    message: str,
    row_number: int | None = None,
    collection: str | None = None,
    preview: dict[str, Any] | None = None,
) -> None:
    dataset_report["warnings"].append(
        {
            "code": code,
            "message": message,
            "rowNumber": row_number,
            "collection": collection,
            "preview": preview or {},
        }
    )


def parse_date(value: Any) -> str | None:
    value = normalize_scalar(value)
    if value is None:
        return None
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None
        if re.fullmatch(r"\d{4}-\d{2}-\d{2}", text):
            return text
        month_match = re.fullmatch(r"(\d{1,2})-([A-Za-z]{3})", text)
        if month_match:
            day = int(month_match.group(1))
            month = datetime.strptime(month_match.group(2), "%b").month
            return date(TODAY.year, month, day).isoformat()
        for fmt in ("%m/%d/%Y", "%m/%d/%y", "%m-%d-%Y", "%Y/%m/%d", "%d-%b-%Y"):
            try:
                parsed = datetime.strptime(text, fmt)
                return date(parsed.year, parsed.month, parsed.day).isoformat()
            except ValueError:
                continue
    return str(value)


def parse_datetime(value: Any) -> str | None:
    value = normalize_scalar(value)
    if value is None:
        return None
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None
        if "T" in text:
            return text
        date_only = parse_date(text)
        if date_only:
            return f"{date_only}T00:00:00"
    return str(value)


def parse_int(value: Any) -> int | None:
    value = normalize_scalar(value)
    if value in (None, ""):
        return None
    try:
        return int(float(str(value).replace(",", "")))
    except ValueError:
        return None


def parse_float(value: Any) -> float | None:
    value = normalize_scalar(value)
    if value in (None, ""):
        return None
    try:
        return round(float(str(value).replace(",", "")), 2)
    except ValueError:
        return None


def parse_currency(value: Any) -> float | None:
    if value in (None, ""):
        return None
    cleaned = str(value).replace("$", "").replace(",", "").strip()
    try:
        return round(float(cleaned), 2)
    except ValueError:
        return None


def parse_yes_no(value: Any) -> bool | None:
    value = normalize_scalar(value)
    if value in (None, ""):
        return None
    text = str(value).strip().lower()
    if text in {"yes", "y", "true", "1", "active", "clear"}:
        return True
    if text in {"no", "n", "false", "0"}:
        return False
    return None


def set_nested(target: dict[str, Any], field: str, value: Any) -> None:
    if value is None:
        return
    cursor = target
    parts = field.split(".")
    for part in parts[:-1]:
        cursor = cursor.setdefault(part, {})
    cursor[parts[-1]] = value


def get_nested(target: dict[str, Any], field: str) -> Any:
    cursor: Any = target
    for part in field.split("."):
        if not isinstance(cursor, dict) or part not in cursor:
            return None
        cursor = cursor[part]
    return cursor


def compose_vehicle_name(values: Iterable[Any]) -> str | None:
    parts = [str(value).strip() for value in values if normalize_scalar(value) not in (None, "")]
    return " ".join(parts) if parts else None


def match_permit_template(value: Any) -> str | None:
    if not value:
        return None
    text = str(value).strip().lower()
    mapping = {
        "state hazmat": "state_hazmat",
        "federal hazmat": "federal_hazmat",
        "ucr": "ucr",
        "operating authority mc150": "operating_authority_mc150",
        "ifta": "ifta",
        "irp": "irp",
    }
    return mapping.get(text, slugify(text))


def infer_reference_contact_category(values: Iterable[Any]) -> str | None:
    joined = " ".join(str(value).lower() for value in values if normalize_scalar(value))
    if "fire" in joined or "emergency" in joined:
        return "emergency"
    if "permit" in joined or "compliance" in joined:
        return "compliance"
    if "environment" in joined or "state" in joined:
        return "regulatory"
    return "operations"


def transform_value(transform: str, raw_value: Any, row: dict[str, Any], source_columns: tuple[str, ...] = ()) -> Any:
    if transform == "identity":
        return normalize_scalar(raw_value)
    if transform in {"normalize_date", "normalize_date_preferred", "normalize_date_secondary"}:
        return parse_date(raw_value)
    if transform == "normalize_datetime":
        return parse_datetime(raw_value)
    if transform == "normalize_time":
        value = normalize_scalar(raw_value)
        if value is None:
            return None
        if isinstance(raw_value, time):
            return raw_value.strftime("%H:%M")
        text = str(value).strip()
        for fmt in ("%H:%M", "%I:%M %p"):
            try:
                return datetime.strptime(text, fmt).strftime("%H:%M")
            except ValueError:
                continue
        return text
    if transform == "normalize_int":
        return parse_int(raw_value)
    if transform == "normalize_float":
        return parse_float(raw_value)
    if transform == "normalize_currency":
        return parse_currency(raw_value)
    if transform == "normalize_yes_no":
        return parse_yes_no(raw_value)
    if transform == "lowercase_email":
        value = normalize_scalar(raw_value)
        return str(value).lower() if value else None
    if transform == "normalize_url":
        value = normalize_scalar(raw_value)
        if not value:
            return None
        text = str(value)
        if text.startswith("http://") or text.startswith("https://"):
            return text
        return f"https://{text}"
    if transform == "uppercase":
        value = normalize_scalar(raw_value)
        return str(value).upper() if value else None
    if transform == "split_csv":
        value = normalize_scalar(raw_value)
        if not value:
            return None
        return [part.strip() for part in str(value).split(",") if part.strip()]
    if transform == "normalize_vehicle_category":
        value = normalize_scalar(raw_value)
        if not value:
            return None
        text = str(value).strip().lower()
        if "truck" in text or "vehicle" in text:
            return "vehicle"
        return "equipment"
    if transform == "normalize_date_or_text":
        return parse_date(raw_value) or normalize_scalar(raw_value)
    if transform == "concat_name":
        first = normalize_scalar(row.get("First Name"))
        last = normalize_scalar(row.get("Last Name"))
        return " ".join(part for part in [first, last] if part) or None
    if transform == "coalesce_date":
        for key in source_columns:
          candidate = parse_date(row.get(key))
          if candidate:
              return candidate
        return None
    if transform == "slug_concat":
        return slugify(*(row.get(column) for column in source_columns))
    if transform == "single_alert_day":
        return parse_int(raw_value)
    if transform == "infer_reference_contact_category":
        return infer_reference_contact_category(row.get(column) for column in source_columns)
    if transform == "compose_vehicle_name":
        return compose_vehicle_name(row.get(column) for column in source_columns)
    if transform == "match_permit_template":
        return match_permit_template(raw_value)
    if transform.startswith("prefix:"):
        prefix = transform.split(":", 1)[1]
        value = normalize_scalar(raw_value)
        return f"{prefix}{slugify(value) or str(value).strip()}" if value else None
    if transform.startswith("prefix_label:"):
        prefix = transform.split(":", 1)[1]
        value = normalize_scalar(raw_value)
        return f"{prefix}{value}" if value else None
    if transform.startswith("literal:"):
        return transform.split(":", 1)[1]
    return normalize_scalar(raw_value)


def read_csv_rows(path: Path) -> list[dict[str, Any]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def read_worksheet_rows(path: Path, worksheet: str) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook[worksheet]
        iterator = sheet.iter_rows(values_only=True)
        headers = [str(value).strip() if value is not None else "" for value in next(iterator)]
        rows = []
        for values in iterator:
            row = {headers[idx]: values[idx] if idx < len(values) else None for idx in range(len(headers))}
            if any(normalize_scalar(value) not in (None, "") for value in row.values()):
                rows.append(row)
        return rows
    finally:
        workbook.close()


def read_config_rows(path: Path, worksheet: str) -> list[dict[str, Any]]:
    return read_worksheet_rows(path, worksheet)


def collection_key_field(dataset, collection: str) -> str | None:
    if not dataset.merge_key:
        return None
    key_collection, key_field = dataset.merge_key.split(".", 1)
    return key_field if key_collection == collection else None


def required_fields_for_dataset_collection(dataset, collection: str) -> list[str]:
    fields: list[str] = []
    for column_mapping in dataset.column_mappings:
        for assignment in column_mapping.assignments:
            if assignment.collection == collection and assignment.required:
                fields.append(assignment.field)
    for derived in dataset.derived_mappings:
        if derived.collection == collection and derived.required:
            fields.append(derived.field)
    return fields


def unique_preserving_order(values: Iterable[str]) -> list[str]:
    seen: set[str] = set()
    ordered: list[str] = []
    for value in values:
        if value in seen:
            continue
        seen.add(value)
        ordered.append(value)
    return ordered


def template_sheet_name(dataset) -> str:
    if dataset.worksheet:
        name = dataset.worksheet
    else:
        name = Path(dataset.source_path).stem
    name = re.sub(r"[\[\]\*\?\/\\:]+", " ", name).strip()
    return name[:31] or "Template"


def template_headers_for_dataset(dataset) -> list[str]:
    headers = [mapping.source_column for mapping in dataset.column_mappings if mapping.source_column and not mapping.ignore]
    return unique_preserving_order(headers)


def template_rows_for_dataset(dataset) -> list[dict[str, Any]]:
    headers = template_headers_for_dataset(dataset)
    if dataset.kind == "config_sheet":
        return [{"Setting": entry.setting_key, "Value": ""} for entry in dataset.config_entry_mappings]
    return [dict.fromkeys(headers, "")]


def build_upload_template_manifest() -> list[dict[str, Any]]:
    manifest: list[dict[str, Any]] = []
    used_names: set[str] = set()
    for dataset in DATASET_MAPPINGS:
        base_name = template_sheet_name(dataset)
        sheet_name = base_name
        suffix = 2
        while sheet_name in used_names:
            trimmed = base_name[: max(0, 31 - len(f" {suffix}"))].rstrip()
            sheet_name = f"{trimmed} {suffix}".strip()
            suffix += 1
        used_names.add(sheet_name)
        manifest.append(
            {
                "datasetKey": dataset.key,
                "kind": dataset.kind,
                "sourcePath": dataset.source_path,
                "worksheet": dataset.worksheet,
                "sheetName": sheet_name,
                "notes": dataset.notes,
                "headers": template_headers_for_dataset(dataset),
                "sampleRows": template_rows_for_dataset(dataset),
            }
        )
    return manifest


def row_has_real_data(record: dict[str, Any]) -> bool:
    for key, value in record.items():
        if key == "orgId":
            continue
        if isinstance(value, dict):
            if row_has_real_data(value):
                return True
            continue
        if value not in (None, "", [], {}):
            return True
    return False


def has_identifier(record: dict[str, Any]) -> bool:
    for key, value in record.items():
        if key == "orgId":
            continue
        if isinstance(value, dict):
            if has_identifier(value):
                return True
            continue
        if key.lower().endswith("id") and value not in (None, "", []):
            return True
    return False


def merge_record(target: dict[str, Any], incoming: dict[str, Any]) -> None:
    for key, value in incoming.items():
        if isinstance(value, dict):
            child = target.setdefault(key, {})
            merge_record(child, value)
        elif value not in (None, "", []):
            target[key] = value


def records_conflict(existing: dict[str, Any], incoming: dict[str, Any]) -> bool:
    for key, value in incoming.items():
        if isinstance(value, dict):
            nested_existing = existing.get(key)
            if isinstance(nested_existing, dict) and records_conflict(nested_existing, value):
                return True
            continue
        existing_value = existing.get(key)
        if existing_value not in (None, "", []) and value not in (None, "", []) and existing_value != value:
            return True
    return False


def finalize_dataset_report(dataset_report: dict[str, Any]) -> dict[str, Any]:
    return {
        **dataset_report,
        "collectionWrites": dict(sorted(dataset_report["collectionWrites"].items())),
    }


def process_dataset(dataset, collections: dict[str, dict[str, dict[str, Any]]], dataset_report: dict[str, Any]) -> None:
    path = ROOT / dataset.source_path
    if dataset.kind == "csv":
        rows = read_csv_rows(path)
    elif dataset.kind == "worksheet":
        rows = read_worksheet_rows(path, dataset.worksheet or "")
    else:
        rows = read_config_rows(path, dataset.worksheet or "")
    dataset_report["rowCount"] = len(rows)

    if dataset.kind == "config_sheet":
        config_map = {str(row.get("Setting") or "").strip(): row.get("Value") for row in rows}
        dataset_report["nonEmptyRows"] = len(config_map)
        org_record = collections["organizations"].setdefault(ORG_ID, {"orgId": ORG_ID})
        for mapping in dataset.config_entry_mappings:
            raw_value = config_map.get(mapping.setting_key)
            value = transform_value(mapping.assignment.transform, raw_value, config_map)
            set_nested(org_record, mapping.assignment.field, value)
        for derived in dataset.derived_mappings:
            value = transform_value(derived.transform, None, config_map, derived.source_columns)
            set_nested(org_record, derived.field, value)
        dataset_report["importedRows"] = 1 if row_has_real_data(org_record) else 0
        dataset_report["collectionWrites"]["organizations"] += 1 if row_has_real_data(org_record) else 0
        return

    for row_index, row in enumerate(rows, start=2):
        if not any(normalize_scalar(value) not in (None, "") for value in row.values()):
            dataset_report["skippedEmptyRows"] += 1
            continue
        dataset_report["nonEmptyRows"] += 1
        staged: dict[str, dict[str, Any]] = defaultdict(dict)
        for column_mapping in dataset.column_mappings:
            raw_value = row.get(column_mapping.source_column)
            if normalize_scalar(raw_value) in (None, ""):
                continue
            for assignment in column_mapping.assignments:
                value = transform_value(assignment.transform, raw_value, row)
                if value is not None:
                    set_nested(staged[assignment.collection], assignment.field, value)

        for derived in dataset.derived_mappings:
            if derived.source_columns and not any(normalize_scalar(row.get(column)) not in (None, "") for column in derived.source_columns):
                if not derived.transform.startswith("literal:"):
                    continue
            sample_value = row.get(derived.source_columns[0]) if derived.source_columns else None
            value = transform_value(derived.transform, sample_value, row, derived.source_columns)
            if value is not None:
                set_nested(staged[derived.collection], derived.field, value)

        wrote_source_row = False
        for collection, record in staged.items():
            if "orgId" not in record:
                record["orgId"] = ORG_ID
            required_fields = required_fields_for_dataset_collection(dataset, collection)
            missing_required = [field for field in required_fields if get_nested(record, field) in (None, "", [])]
            if missing_required:
                dataset_report["skippedRows"] += 1
                add_dataset_warning(
                    dataset_report,
                    code="missing_required_fields",
                    message=f"Skipped {collection} row because required fields were empty: {', '.join(missing_required)}.",
                    row_number=row_index,
                    collection=collection,
                    preview=source_row_preview(row),
                )
                continue
            if not row_has_real_data(record):
                continue
            key_field = collection_key_field(dataset, collection)
            key_value = get_nested(record, key_field) if key_field else None
            if key_field and not key_value:
                key_value = stable_row_id(dataset.key, collection, row)
                set_nested(record, key_field, key_value)
                dataset_report["generatedIds"].append(
                    {
                        "collection": collection,
                        "field": key_field,
                        "rowNumber": row_index,
                        "generatedId": key_value,
                        "reason": "missing_merge_key_field",
                        "preview": source_row_preview(row),
                    }
                )
            if not key_field and not has_identifier(record):
                dataset_report["skippedRows"] += 1
                add_dataset_warning(
                    dataset_report,
                    code="missing_identifier",
                    message=f"Skipped {collection} row because it had no stable identifier fields.",
                    row_number=row_index,
                    collection=collection,
                    preview=source_row_preview(row),
                )
                continue
            if not key_value:
                key_value = stable_row_id(dataset.key, collection, row)
                if key_field:
                    set_nested(record, key_field, key_value)
            storage_key = str(key_value)
            existing = collections[collection].get(storage_key)
            if existing and records_conflict(existing, record):
                collision_key = f"{storage_key}-{stable_hash(dataset.key, collection, row)}"
                if key_field:
                    set_nested(record, key_field, collision_key)
                storage_key = collision_key
                dataset_report["generatedIds"].append(
                    {
                        "collection": collection,
                        "field": key_field or "_storageKey",
                        "rowNumber": row_index,
                        "generatedId": storage_key,
                        "reason": "merge_key_collision",
                        "preview": source_row_preview(row),
                    }
                )
                add_dataset_warning(
                    dataset_report,
                    code="merge_key_collision",
                    message=f"Duplicate merge key detected for {collection}; a collision-safe key was generated.",
                    row_number=row_index,
                    collection=collection,
                    preview=source_row_preview(row),
                )
            existing = collections[collection].setdefault(storage_key, {})
            merge_record(existing, record)
            dataset_report["collectionWrites"][collection] += 1
            wrote_source_row = True
        if wrote_source_row:
            dataset_report["importedRows"] += 1


def compact_collection_rows(rows: dict[str, dict[str, Any]]) -> list[dict[str, Any]]:
    output = []
    for record in rows.values():
        if row_has_real_data(record):
            output.append(record)
    return sorted(output, key=lambda record: json.dumps(record, sort_keys=True))


def seed_fallbacks(collections: dict[str, list[dict[str, Any]]]) -> list[str]:
    fallbacks_used: list[str] = []
    if not collections["permit_license_records"]:
        fallbacks_used.append("permit_license_records")
        collections["permit_license_records"] = [
            {
                "recordId": "permit_state_hazmat_2026",
                "orgId": ORG_ID,
                "templateId": "state_hazmat",
                "entityType": "organization",
                "entityId": ORG_ID,
                "permitNumber": "CO-HM-2026-001",
                "issuingAgency": "Colorado Department of Public Health & Environment",
                "issueDate": "2026-01-01",
                "expirationDate": "2026-12-31",
                "renewalDueDate": "2026-12-01",
                "status": "active",
                "responsibleEmployeeName": "Fleet Manager",
                "notes": "Seeded because permits CSV is currently blank.",
            },
            {
                "recordId": "permit_ifta_2026_q2",
                "orgId": ORG_ID,
                "templateId": "ifta",
                "entityType": "organization",
                "entityId": ORG_ID,
                "permitNumber": "IFTA-2026-Q2",
                "issuingAgency": "Colorado Department of Revenue",
                "issueDate": "2026-04-01",
                "expirationDate": "2026-06-30",
                "renewalDueDate": "2026-06-30",
                "status": "active",
                "responsibleEmployeeName": "Compliance Coordinator",
                "notes": "Seeded because permits CSV is currently blank.",
            },
            {
                "recordId": "permit_mc150_2026",
                "orgId": ORG_ID,
                "templateId": "operating_authority_mc150",
                "entityType": "organization",
                "entityId": ORG_ID,
                "permitNumber": "MC150-2026",
                "issuingAgency": "FMCSA",
                "issueDate": "2026-01-01",
                "expirationDate": "2026-10-31",
                "renewalDueDate": "2026-10-31",
                "status": "planned",
                "responsibleEmployeeName": "Operations Manager",
                "notes": "Seeded compliance milestone until source row exists.",
            },
        ]

    if not collections["tank_assets"]:
        fallbacks_used.append("tank_assets")
        collections["tank_assets"] = [
            {
                "assetId": "tank_t001",
                "orgId": ORG_ID,
                "tankId": "T-001",
                "location": "North Yard",
                "tankType": "UST",
                "capacityGallons": 10000,
                "productStored": "ULSD",
                "installationDate": "2021-06-01",
                "ustPermitNumber": "UST-2024-001",
                "permitExpiration": "2026-12-31",
                "lastInspectionDate": "2026-01-15",
                "nextInspectionDue": "2027-01-15",
                "lastLeakTest": "2026-01-15",
                "nextLeakTestDue": "2026-07-15",
                "status": "active",
                "notes": "Seeded tank record until storage tank CSV is populated.",
            }
        ]

    asset_ids = {row.get("assetId") for row in collections["assets"]}
    for tank in collections["tank_assets"]:
        asset_id = tank.get("assetId")
        if asset_id in asset_ids:
            continue
        collections["assets"].append(
            {
                "assetId": asset_id,
                "orgId": ORG_ID,
                "category": "tank",
                "name": f"Tank {tank.get('tankId')}",
                "status": tank.get("status", "active"),
                "location": tank.get("location"),
                "notes": tank.get("notes"),
            }
        )
        asset_ids.add(asset_id)
    return fallbacks_used


def generate_suspense_items(collections: dict[str, list[dict[str, Any]]]) -> list[dict[str, Any]]:
    items: list[dict[str, Any]] = []

    def add_item(source_type: str, source_id: str, title: str, due_date: str | None, owner_email: str, severity: str) -> None:
        if not due_date:
            return
        due = datetime.strptime(due_date, "%Y-%m-%d").date()
        delta = (due - TODAY).days
        if delta < 0:
            alert_state = "overdue"
        elif delta <= 7:
            alert_state = "7_day"
        elif delta <= 30:
            alert_state = "30_day"
        else:
            alert_state = "scheduled"
        items.append(
            {
                "suspenseItemId": slugify(source_type, source_id, due_date) or f"{source_type}-{source_id}",
                "orgId": ORG_ID,
                "sourceType": source_type,
                "sourceId": source_id,
                "templateId": None,
                "title": title,
                "ownerEmail": owner_email,
                "severity": severity,
                "status": "open",
                "dueDate": due_date,
                "alertState": alert_state,
                "lastAlertSentAt": None,
            }
        )

    for row in collections["employee_compliance"]:
        person_id = row.get("personId") or row.get("employeeId") or row.get("driverId") or "driver"
        owner = "fleet@samplefleet.example"
        if row.get("medicalExpiration"):
            add_item("employee_compliance", str(person_id), f"Medical card expiration for {person_id}", row["medicalExpiration"], owner, "high")
        if row.get("nextMvrDue"):
            add_item("employee_compliance", str(person_id), f"MVR review due for {person_id}", row["nextMvrDue"], owner, "medium")
        if row.get("tsaExpiration"):
            add_item("employee_compliance", str(person_id), f"TSA / HazMat threat assessment renewal for {person_id}", row["tsaExpiration"], "compliance@samplefleet.example", "medium")

    for row in collections["permit_license_records"]:
        due_date = row.get("renewalDueDate") or row.get("expirationDate")
        owner = "compliance@samplefleet.example"
        title = f"{row.get('templateId') or row.get('permitType') or 'Permit'} renewal due"
        add_item("permit_license_records", str(row.get("recordId")), title, due_date, owner, "medium")

    for row in collections["assets"]:
        if row.get("nextServiceDue"):
            add_item("assets", str(row.get("assetId")), f"Service due for {row.get('name')}", row["nextServiceDue"], "maintenance@samplefleet.example", "medium")

    for row in collections["tank_assets"]:
        if row.get("nextLeakTestDue"):
            add_item("tank_assets", str(row.get("tankId") or row.get("assetId")), f"Leak test due for tank {row.get('tankId')}", row["nextLeakTestDue"], "operations@samplefleet.example", "high")

    deduped = {item["suspenseItemId"]: item for item in items}
    return sorted(deduped.values(), key=lambda item: item["dueDate"])


def build_imports() -> tuple[dict[str, list[dict[str, Any]]], dict[str, Any], list[dict[str, Any]]]:
    staging: dict[str, dict[str, dict[str, Any]]] = defaultdict(dict)
    dataset_reports: list[dict[str, Any]] = []
    for dataset in DATASET_MAPPINGS:
        dataset_report = make_dataset_report(dataset)
        process_dataset(dataset, staging, dataset_report)
        dataset_reports.append(finalize_dataset_report(dataset_report))

    collections: dict[str, list[dict[str, Any]]] = {collection: compact_collection_rows(rows) for collection, rows in staging.items()}
    for name in [
        "organizations",
        "people",
        "employee_compliance",
        "emergency_contacts",
        "assets",
        "vehicle_assets",
        "tank_assets",
        "activity_logs",
        "maintenance_plans",
        "maintenance_events",
        "permit_license_records",
        "reference_contacts",
    ]:
        collections.setdefault(name, [])

    fallbacks_used = seed_fallbacks(collections)
    for dataset_report in dataset_reports:
        if dataset_report["datasetKey"] == "csv:PERMITS & LICENSES.csv" and "permit_license_records" in fallbacks_used:
            dataset_report["fallbacksUsed"].append("permit_license_records")
        if dataset_report["datasetKey"] == "csv:STORAGE TANKS.csv" and "tank_assets" in fallbacks_used:
            dataset_report["fallbacksUsed"].append("tank_assets")
    collections["suspense_items"] = generate_suspense_items(collections)
    upload_template_manifest = build_upload_template_manifest()
    validation_report = {
        "generatedAt": f"{TODAY.isoformat()}T00:00:00",
        "orgId": ORG_ID,
        "summary": {
            "collectionCounts": {name: len(rows) for name, rows in sorted(collections.items())},
            "datasetCount": len(dataset_reports),
            "warningCount": sum(len(dataset["warnings"]) for dataset in dataset_reports),
            "generatedIdCount": sum(len(dataset["generatedIds"]) for dataset in dataset_reports),
            "fallbackCollections": fallbacks_used,
        },
        "datasets": dataset_reports,
    }
    return collections, validation_report, upload_template_manifest


def write_json_outputs(
    collections: dict[str, list[dict[str, Any]]],
    validation_report: dict[str, Any],
    upload_template_manifest: list[dict[str, Any]],
) -> None:
    GENERATED_DIR.mkdir(exist_ok=True)
    for name, rows in collections.items():
        path = GENERATED_DIR / f"{name}.json"
        path.write_text(json.dumps(rows, indent=2), encoding="utf-8")
    GENERATED_VALIDATION_REPORT_PATH.write_text(json.dumps(validation_report, indent=2), encoding="utf-8")
    GENERATED_TEMPLATE_MANIFEST_PATH.write_text(json.dumps(upload_template_manifest, indent=2), encoding="utf-8")


def to_ts(value: Any) -> str:
    return json.dumps(value, indent=2)


def write_ts_module(
    collections: dict[str, list[dict[str, Any]]],
    upload_template_manifest: list[dict[str, Any]],
) -> None:
    GENERATED_TS_PATH.parent.mkdir(parents=True, exist_ok=True)
    lines = [
        "/* auto-generated by tooling/chief-sentinel/build_chief_imports.py */",
        "",
    ]
    for name, rows in sorted(collections.items()):
        export_name = re.sub(r"[^a-zA-Z0-9]", "_", name)
        lines.append(f"export const {export_name} = {to_ts(rows)} as const;")
        lines.append("")
    GENERATED_TS_PATH.write_text("\n".join(lines), encoding="utf-8")
    GENERATED_TEMPLATE_TS_PATH.write_text(
        "\n".join(
            [
                "/* auto-generated by tooling/chief-sentinel/build_chief_imports.py */",
                "",
                f"export const chiefUploadTemplateManifest = {to_ts(upload_template_manifest)} as const;",
                "",
            ]
        ),
        encoding="utf-8",
    )


def write_validation_report_markdown(validation_report: dict[str, Any]) -> None:
    lines = [
        "# Chief Import Validation Report",
        "",
        f"Generated: `{validation_report['generatedAt']}`",
        "",
        "## Summary",
        "",
    ]
    for name, count in validation_report["summary"]["collectionCounts"].items():
        lines.append(f"- `{name}`: {count}")
    lines.extend(
        [
            f"- `warningCount`: {validation_report['summary']['warningCount']}",
            f"- `generatedIdCount`: {validation_report['summary']['generatedIdCount']}",
            f"- `fallbackCollections`: {', '.join(validation_report['summary']['fallbackCollections']) or 'none'}",
            "",
            "## Dataset Notes",
            "",
        ]
    )
    for dataset in validation_report["datasets"]:
        lines.append(f"### {dataset['datasetKey']}")
        lines.append("")
        lines.append(f"- `rows`: {dataset['rowCount']}")
        lines.append(f"- `importedRows`: {dataset['importedRows']}")
        lines.append(f"- `collectionWrites`: {json.dumps(dataset['collectionWrites'], sort_keys=True)}")
        if dataset["fallbacksUsed"]:
            lines.append(f"- `fallbacksUsed`: {', '.join(dataset['fallbacksUsed'])}")
        if dataset["generatedIds"]:
            lines.append(f"- `generatedIds`: {len(dataset['generatedIds'])}")
        if dataset["warnings"]:
            lines.append(f"- `warnings`: {len(dataset['warnings'])}")
        lines.append("")
    GENERATED_VALIDATION_REPORT_MD_PATH.write_text("\n".join(lines), encoding="utf-8")


def print_summary(collections: dict[str, list[dict[str, Any]]], validation_report: dict[str, Any]) -> None:
    print("Generated imports:")
    for name in sorted(collections.keys()):
        print(f"  {name}: {len(collections[name])}")
    print(f"\nJSON output: {GENERATED_DIR}")
    print(f"TS module: {GENERATED_TS_PATH}")
    print(f"Template manifest: {GENERATED_TEMPLATE_MANIFEST_PATH}")
    print(f"Validation report: {GENERATED_VALIDATION_REPORT_PATH}")
    print(f"Warnings: {validation_report['summary']['warningCount']}")
    print(f"Generated fallback IDs: {validation_report['summary']['generatedIdCount']}")


def main() -> int:
    collections, validation_report, upload_template_manifest = build_imports()
    write_json_outputs(collections, validation_report, upload_template_manifest)
    write_ts_module(collections, upload_template_manifest)
    write_validation_report_markdown(validation_report)
    print_summary(collections, validation_report)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

